"""Form structure parsing and smart insertion logic."""

# Ensure UTF-8 encoding for Windows console output
import sys
if sys.platform == 'win32':
    import locale
    import codecs
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except AttributeError:
        # Python < 3.7 doesn't have reconfigure, use environment variable approach
        import os
        os.environ['PYTHONIOENCODING'] = 'utf-8'

from typing import List, Optional, Dict


# Metadata field types
METADATA_TYPES = {
    'start', 'end', 'today', 'deviceid', 'subscriberid',
    'simserial', 'phonenumber', 'username', 'email', 'audit'
}

def _cell_has_value(value) -> bool:
    """Return True if the cell value is meaningful (not just whitespace)."""
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _row_has_data(ws, row_idx: int, check_columns: List[int]) -> bool:
    """Check if a row has meaningful data in any of the specified columns."""
    for col_idx in check_columns:
        cell_value = ws.cell(row_idx, col_idx).value
        if _cell_has_value(cell_value):
            return True
    return False


def is_metadata_field(question_type: str) -> bool:
    """Check if a question type is a metadata field.

# CRITICAL: Add scripts directory to Python path for sibling imports
# This allows the script to find sibling modules whether run from project root or scripts dir
_scripts_dir = Path(__file__).parent.resolve()
if str(_scripts_dir) not in sys.path:
    sys.path.insert(0, str(_scripts_dir))

    Args:
        question_type: XLSForm question type

    Returns:
        True if the question type is a metadata field
    """
    if not question_type:
        return False
    return question_type.lower() in METADATA_TYPES


def find_last_data_row(ws, header_row: int = 1, check_columns: List[int] = None) -> int:
    """
    Find the last row that contains actual data (not just formatting).

    This is needed when templates have pre-formatted empty rows.
    Using ws.max_row directly would return the last formatted row,
    not the last row with data.

    Args:
        ws: openpyxl worksheet object
        header_row: Row number of the header (default: 1)
        check_columns: List of column indices to check (1-indexed).
                      If None, checks columns 1-3.

    Returns:
        The row number of the last data row, or header_row if no data found
    """
    if check_columns is None:
        check_columns = [1, 2, 3]  # Check first 3 columns by default

    last_data_row = header_row

    for row_idx in range(header_row + 1, ws.max_row + 1):
        if _row_has_data(ws, row_idx, check_columns):
            last_data_row = row_idx

    return last_data_row


def find_insertion_point(ws, header_row: int, questions_data: List[dict], column_map: Dict[str, int] = None) -> int:
    """
    Determine where to insert new questions.

    Strategy:
    1. If adding metadata: insert immediately after header row (before any questions)
    2. If adding regular questions: insert after last metadata or last question

    Args:
        ws: openpyxl worksheet object
        header_row: Row number of the header
        questions_data: List of question dicts to insert
        column_map: Optional dictionary mapping column names to positions.
                    If not provided, assumes type=1, name=2.

    Returns:
        The row number where insertion should begin
    """
    # Default column mapping if not provided (for backward compatibility)
    if column_map is None:
        column_map = {"type": 1, "name": 2}

    # Get column positions (with fallbacks)
    type_col = column_map.get("type", 1)
    name_col = column_map.get("name", 2)
    check_columns = [type_col, name_col]

    # Check if we're adding metadata fields
    has_metadata = any(is_metadata_field(q.get('type', '')) for q in questions_data)

    if has_metadata:
        # Find end of existing metadata section
        # Insert before first non-metadata question
        last_data_row = find_last_data_row(ws, header_row, check_columns)
        for row_idx in range(header_row + 1, last_data_row + 1):
            cell_type = ws.cell(row_idx, type_col).value
            cell_type = str(cell_type).strip() if cell_type is not None else ""
            # Stop at first empty row - this is where we should insert
            if not cell_type:
                return row_idx
            # If we find a non-metadata field, insert here
            if not is_metadata_field(str(cell_type)):
                return row_idx
        # No empty rows or non-metadata questions found, append at end of data
        return last_data_row + 1
    else:
        # Find last question (after any metadata) based on actual data
        last_data_row = find_last_data_row(ws, header_row, check_columns)
        return last_data_row + 1


def freeze_top_row(ws, header_row: int = 1):
    """
    Freeze the top row (header) in Excel so it stays visible when scrolling.

    Args:
        ws: openpyxl worksheet object
        header_row: Row number to freeze below (default: 1)
    """
    # openpyxl uses 1-based indexing
    # Freeze panes at the row below the header
    ws.freeze_panes = f"A{header_row + 1}"


def find_header_row(ws) -> Optional[int]:
    """
    Find the header row by looking for the 'type' column.

    Searches across multiple columns (not just column 1) to handle
    XLSForm templates with columns in non-standard positions.

    Args:
        ws: openpyxl worksheet object

    Returns:
        Row number of the header, or None if not found
    """
    # Check first 10 rows for header
    for row_idx in range(1, min(10, ws.max_row + 1)):
        # Search across first 20 columns for 'type' (more robust than just column 1)
        max_col = min(ws.max_column, 20)
        for col_idx in range(1, max_col + 1):
            cell_value = ws.cell(row_idx, col_idx).value
            if cell_value and str(cell_value).strip().lower() == "type":
                return row_idx
    return None


def build_column_mapping(ws, header_row: int) -> Dict[str, int]:
    """
    Build a dynamic mapping of column names to column positions.

    Reads the header row to determine where each column actually is,
    instead of relying on hardcoded positions. This allows the system
    to work with XLSForm templates that have columns in any order.

    Args:
        ws: openpyxl worksheet object
        header_row: Row number containing headers (1-indexed)

    Returns:
        Dictionary mapping column names to column numbers (1-indexed)
        Example: {"type": 1, "name": 3, "label": 5}

    Note:
        Returns empty dict if header_row is invalid or no headers found.
        Column names are converted to lowercase and stripped of whitespace
        for case-insensitive matching.
    """
    column_map = {}

    # Iterate through all columns in the header row
    # openpyxl's max_column may not always be accurate, so we check a reasonable range
    max_col = min(ws.max_column, 100)  # Sanity check: don't scan more than 100 columns

    for col_idx in range(1, max_col + 1):
        cell_value = ws.cell(header_row, col_idx).value
        if cell_value:
            # Convert to string, strip whitespace, and convert to lowercase
            # for consistent, case-insensitive lookups
            col_name = str(cell_value).strip().lower()
            if col_name:
                column_map[col_name] = col_idx

    return column_map


class FormStructure:
    """Parse and understand XLSForm structure for context-aware insertion."""

    def __init__(self, worksheet, header_row, column_map: Dict[str, int] = None):
        """Initialize form structure parser.

        Args:
            worksheet: openpyxl worksheet object
            header_row: Row number of the header
            column_map: Optional dictionary mapping column names to positions.
                        If not provided, assumes type=1, name=2.
        """
        self.ws = worksheet
        self.header_row = header_row
        # Default column mapping if not provided (for backward compatibility)
        self.column_map = column_map if column_map is not None else {"type": 1, "name": 2}
        self.structure = self._parse_structure()

    def _parse_structure(self) -> dict:
        """Parse form to identify groups, repeats, and questions.

        Returns:
            Dictionary with keys: metadata, groups, repeats, questions
        """
        structure = {
            'metadata': [],
            'groups': [],
            'repeats': [],
            'questions': []
        }

        # Get column positions from the column map
        type_col = self.column_map.get("type", 1)
        name_col = self.column_map.get("name", 2)

        stack = []  # Track current context (group/repeat)

        for row_idx in range(self.header_row + 1, self.ws.max_row + 1):
            cell_type = str(self.ws.cell(row_idx, type_col).value or '').strip()
            cell_name = str(self.ws.cell(row_idx, name_col).value or '').strip()

            if not cell_type:
                continue

            # Handle group/repeat boundaries
            if cell_type == 'begin group':
                stack.append(('group', row_idx, cell_name))
            elif cell_type == 'end group':
                if stack and stack[-1][0] == 'group':
                    _, start_row, name = stack.pop()
                    structure['groups'].append({
                        'name': name, 'start_row': start_row, 'end_row': row_idx
                    })
            elif cell_type == 'begin repeat':
                stack.append(('repeat', row_idx, cell_name))
            elif cell_type == 'end repeat':
                if stack and stack[-1][0] == 'repeat':
                    _, start_row, name = stack.pop()
                    structure['repeats'].append({
                        'name': name, 'start_row': start_row, 'end_row': row_idx
                    })

            # Track questions
            elif is_metadata_field(cell_type):
                structure['metadata'].append({
                    'row': row_idx, 'type': cell_type, 'name': cell_name
                })
            elif cell_type not in ['begin group', 'end group', 'begin repeat', 'end repeat']:
                structure['questions'].append({
                    'row': row_idx, 'type': cell_type, 'name': cell_name
                })

        return structure

    def get_insertion_point_in_context(self, context_name: str) -> Optional[int]:
        """
        Get row number to insert question within a group/repeat.

        Args:
            context_name: Name of the group or repeat

        Returns:
            Row number where insertion should begin, or None if context not found
        """
        # Find the context
        context = None
        for item in self.structure['groups'] + self.structure['repeats']:
            if item['name'] == context_name:
                context = item
                break

        if not context:
            return None

        # Find last question in this context
        last_row = context['start_row']
        for q in self.structure['questions']:
            if context['start_row'] < q['row'] < context['end_row']:
                last_row = q['row']

        return last_row + 1

    def find_question_row(self, question_name: str) -> Optional[int]:
        """
        Find the row number of a specific question.

        Args:
            question_name: Name of the question to find

        Returns:
            Row number, or None if not found
        """
        for q in self.structure['questions']:
            if q['name'] == question_name:
                return q['row']
        return None


def add_visual_separation(ws, insert_row: int, is_new_context: bool = False) -> int:
    """
    Add blank rows for visual separation when needed.

    Args:
        ws: openpyxl worksheet object
        insert_row: Row number where insertion should happen
        is_new_context: Whether this is starting a new group/repeat

    Returns:
        Adjusted row number after potential blank row insertion
    """
    if is_new_context:
        # Insert 2 blank rows for clear visual separation
        ws.insert_rows(insert_row, 2)
        return insert_row + 2
    return insert_row


def ensure_blank_row_gap(ws, insert_row: int, header_row: int, check_columns: List[int], min_blank_rows: int = 1) -> int:
    """
    Ensure a minimum number of blank rows before the insertion point.

    This supports readability by keeping at least one blank row between
    sections/blocks of questions.
    """
    if min_blank_rows <= 0 or insert_row <= header_row + 1:
        return insert_row

    blank_rows = 0
    row_idx = insert_row - 1
    while row_idx > header_row and blank_rows < min_blank_rows:
        if _row_has_data(ws, row_idx, check_columns):
            break
        blank_rows += 1
        row_idx -= 1

    if blank_rows >= min_blank_rows:
        return insert_row

    rows_to_insert = min_blank_rows - blank_rows
    ws.insert_rows(insert_row, rows_to_insert)
    return insert_row + rows_to_insert


if __name__ == "__main__":
    # Test form structure parsing
    import sys
    import openpyxl
    from pathlib import Path

    if len(sys.argv) < 2:
        print("Usage: python form_structure.py <xlsx_file>")
        sys.exit(1)

    xlsx_file = sys.argv[1]

    # Check if file exists
    if not Path(xlsx_file).exists():
        print(f"Error: File '{xlsx_file}' not found")
        sys.exit(1)

    # Check if file has content (not empty)
    if Path(xlsx_file).stat().st_size == 0:
        print(f"Error: File '{xlsx_file}' is empty")
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(xlsx_file)
    except Exception as e:
        print(f"Error: Could not load Excel file: {e}")
        sys.exit(1)

    if "survey" not in wb.sheetnames:
        print("Error: 'survey' sheet not found")
        sys.exit(1)

    ws = wb["survey"]

    # Find header row
    header_row = find_header_row(ws)
    if not header_row:
        print("Error: Could not find header row")
        sys.exit(1)

    print(f"Header row: {header_row}")

    # Parse structure
    form_structure = FormStructure(ws, header_row)

    # Log structure analysis
    try:
        from log_activity import ActivityLogger
        logger = ActivityLogger()
        structure_info = (
            f"Total questions: {len(form_structure.structure['questions'])}\n"
            f"Groups: {len(form_structure.structure['groups'])}\n"
            f"Repeats: {len(form_structure.structure['repeats'])}\n"
            f"Metadata fields: {len(form_structure.structure['metadata'])}"
        )
        logger.log_action(
            action_type="analyze_structure",
            description="Analyzed form structure",
            details=structure_info
        )
    except Exception:
        # Silently fail if logging is not available
        pass

    print(f"\nMetadata fields: {len(form_structure.structure['metadata'])}")
    for m in form_structure.structure['metadata']:
        print(f"  Row {m['row']}: {m['type']} - {m['name']}")

    print(f"\nGroups: {len(form_structure.structure['groups'])}")
    for g in form_structure.structure['groups']:
        print(f"  {g['name']}: rows {g['start_row']}-{g['end_row']}")

    print(f"\nRepeats: {len(form_structure.structure['repeats'])}")
    for r in form_structure.structure['repeats']:
        print(f"  {r['name']}: rows {r['start_row']}-{r['end_row']}")

    print(f"\nQuestions: {len(form_structure.structure['questions'])}")
    for q in form_structure.structure['questions'][:5]:  # Show first 5
        print(f"  Row {q['row']}: {q['type']} - {q['name']}")
    if len(form_structure.structure['questions']) > 5:
        print(f"  ... and {len(form_structure.structure['questions']) - 5} more")
