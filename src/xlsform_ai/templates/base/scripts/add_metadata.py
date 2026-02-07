#!/usr/bin/env python3
"""
Add XLSForm metadata fields to survey sheet.

Standard metadata fields:
- start: Form start timestamp
- end: Form end timestamp
- today: Date of submission
- deviceid: Device identifier
- subscriberid: Subscriber ID
- simserial: SIM serial number
- phonenumber: Phone number
- username: Enumerator username

Usage:
    python add_metadata.py
    python add_metadata.py --file survey.xlsx
"""

import sys
import openpyxl
from pathlib import Path

# Ensure UTF-8 encoding for Windows console output
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except AttributeError:
        pass

# Standard XLSForm metadata fields
METADATA_FIELDS = [
    {"type": "start", "name": "start", "label": "Start"},
    {"type": "end", "name": "end", "label": "End"},
    {"type": "today", "name": "today", "label": "Today"},
    {"type": "deviceid", "name": "deviceid", "label": "Device ID"},
    {"type": "subscriberid", "name": "subscriberid", "label": "Subscriber ID"},
    {"type": "simserial", "name": "simserial", "label": "SIM Serial"},
    {"type": "phonenumber", "name": "phonenumber", "label": "Phone Number"},
    {"type": "username", "name": "username", "label": "Username"},
]

def _cell_has_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def add_metadata_fields(survey_file="survey.xlsx"):
    """Add standard XLSForm metadata fields to the survey sheet.

    Args:
        survey_file: Path to XLSForm file

    Returns:
        dict with success status and details
    """
    try:
        survey_path = Path(survey_file)

        if not survey_path.exists():
            return {"success": False, "error": f"File not found: {survey_file}"}

        # Load workbook
        wb = openpyxl.load_workbook(survey_path)

        if "survey" not in wb.sheetnames:
            return {"success": False, "error": "'survey' sheet not found"}

        ws = wb["survey"]

        # Find header row
        header_row = None
        for row_idx in range(1, min(10, ws.max_row + 1)):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value and str(cell_value).strip().lower() == "type":
                header_row = row_idx
                break

        if header_row is None:
            return {"success": False, "error": "Could not find header row"}

        # Build column mapping
        column_map = {}
        for col_idx in range(1, min(30, ws.max_column + 1)):
            header_val = ws.cell(header_row, col_idx).value
            if header_val:
                column_map[str(header_val).lower()] = col_idx

        required_columns = ["type", "name", "label"]
        missing_columns = [col for col in required_columns if col not in column_map]
        if missing_columns:
            return {"success": False, "error": f"Missing columns: {missing_columns}"}

        # Check for existing metadata fields
        name_col = column_map.get("name", 2)
        existing_metadata = set()
        for row_idx in range(header_row + 1, ws.max_row + 1):
            name_val = ws.cell(row_idx, name_col).value
            if _cell_has_value(name_val):
                cleaned_name = str(name_val).strip().lower()
                if cleaned_name in [f["name"] for f in METADATA_FIELDS]:
                    existing_metadata.add(cleaned_name)

        # Filter out metadata fields that already exist
        to_add = [f for f in METADATA_FIELDS if f["name"] not in existing_metadata]

        if not to_add:
            return {
                "success": True,
                "added": [],
                "skipped": list(existing_metadata),
                "total": 0,
                "message": "All metadata fields already exist"
            }

        # Insert metadata fields at the top (after header row)
        # Shift existing rows down if needed
        first_data_row = header_row + 1
        rows_to_insert = len(to_add)

        # Check if we need to shift existing data
        has_existing_data = False
        for row_idx in range(first_data_row, first_data_row + rows_to_insert):
            if row_idx <= ws.max_row:
                name_val = ws.cell(row_idx, name_col).value
                if _cell_has_value(name_val):
                    has_existing_data = True
                    break

        if has_existing_data:
            # Insert rows to make space
            ws.insert_rows(first_data_row, rows_to_insert)

        # Add metadata fields
        added = []
        for idx, field in enumerate(to_add):
            row_idx = first_data_row + idx
            ws.cell(row_idx, column_map["type"], field["type"])
            ws.cell(row_idx, column_map["name"], field["name"])
            ws.cell(row_idx, column_map["label"], field["label"])
            added.append({
                "row": row_idx,
                "name": field["name"],
                "type": field["type"]
            })

        # Save workbook
        wb.save(survey_path)

        return {
            "success": True,
            "added": added,
            "skipped": list(existing_metadata),
            "total": len(added)
        }

    except Exception as e:
        return {"success": False, "error": str(e)}


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='Add XLSForm metadata fields')
    parser.add_argument('--file', '-f', default='survey.xlsx',
                       help='XLSForm file (default: survey.xlsx)')

    args = parser.parse_args()

    result = add_metadata_fields(args.file)

    if result["success"]:
        added = result.get("added", [])
        skipped = result.get("skipped", [])

        if added:
            print(f"Successfully added {len(added)} metadata field(s):")
            for field in added:
                print(f"  Row {field['row']}: {field['type']} ({field['name']})")

        if skipped:
            print(f"\nSkipped {len(skipped)} existing field(s): {', '.join(skipped)}")

        if not added and skipped:
            print("\nAll metadata fields are already present in the form.")
        sys.exit(0)
    else:
        print(f"Error: {result['error']}", file=sys.stderr)
        sys.exit(1)
