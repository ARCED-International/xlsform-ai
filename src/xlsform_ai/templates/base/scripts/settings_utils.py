"""Utilities for reading and updating XLSForm settings sheet."""

import sys
from pathlib import Path
from typing import Dict, List, Optional

try:
    import openpyxl
except ImportError as exc:
    raise ImportError("openpyxl is required for settings utilities") from exc

try:
    from history_manager import WorkbookHistoryManager
    HISTORY_AVAILABLE = True
except Exception:
    HISTORY_AVAILABLE = False


SETTINGS_SHEET_NAME = "settings"
REQUIRED_SETTINGS_COLUMNS = ["form_title", "form_id"]
PROTECTED_SETTINGS_COLUMNS = {"version"}
VERSION_FORMULA = '=TEXT(NOW(), "yyyymmddhhmmss")'
_KEY_VALUE_HEADERS = {"column_name", "value"}
_KNOWN_SETTINGS_COLUMNS = [
    "form_title",
    "form_id",
    "version",
    "instance_name",
    "default_language",
    "public_key",
    "submission_url",
    "style",
    "name",
    "clean_text_values",
]


def _get_settings_sheet(wb):
    if SETTINGS_SHEET_NAME in wb.sheetnames:
        return wb[SETTINGS_SHEET_NAME]
    return None


def _normalize_header_name(value: str) -> str:
    try:
        from form_structure import normalize_header_name
    except Exception:
        normalize_header_name = None
    if normalize_header_name:
        return normalize_header_name(value)
    return str(value).strip().lower().replace(" ", "_")


def _get_header_map(sheet) -> Dict[str, int]:
    headers = [cell.value for cell in sheet[1]]
    header_map = {}
    for idx, header in enumerate(headers, start=1):
        if header:
            header_map[str(header).strip()] = idx
    return header_map


def _get_header_map_lower(sheet) -> Dict[str, int]:
    headers = [cell.value for cell in sheet[1]]
    header_map = {}
    for idx, header in enumerate(headers, start=1):
        if header:
            header_name = str(header).strip()
            header_key = _normalize_header_name(header_name)
            header_map[header_key] = idx
    return header_map


def _find_settings_header_row(sheet, max_rows: int = 5) -> Optional[int]:
    """Find the header row within the first few rows."""
    for row_idx in range(1, min(max_rows, sheet.max_row) + 1):
        row_values = [cell.value for cell in sheet[row_idx]]
        header_keys = {_normalize_header_name(v) for v in row_values if v}
        if not header_keys:
            continue
        if _KEY_VALUE_HEADERS.issubset(header_keys):
            return row_idx
        if any(key in header_keys for key in _KNOWN_SETTINGS_COLUMNS):
            return row_idx
    return None


def _rebuild_settings_sheet(sheet, headers: List[str], values: Dict[str, object]) -> None:
    """Rewrite settings sheet as row 1 headers, row 2 values."""
    sheet.delete_rows(1, sheet.max_row)
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)
        if header in values:
            sheet.cell(row=2, column=col_idx, value=values[header])


def _extract_pairs_from_key_value(sheet, header_row: int) -> Dict[str, object]:
    header_cells = [cell.value for cell in sheet[header_row]]
    header_keys = {_normalize_header_name(v): idx + 1 for idx, v in enumerate(header_cells) if v}
    name_col = header_keys.get("column_name")
    value_col = header_keys.get("value")
    if not name_col or not value_col:
        return {}
    pairs: Dict[str, object] = {}
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        key = sheet.cell(row=row_idx, column=name_col).value
        value = sheet.cell(row=row_idx, column=value_col).value
        if key is None:
            continue
        key_str = str(key).strip()
        if not key_str:
            continue
        pairs[_normalize_header_name(key_str)] = value
    return pairs


def _extract_pairs_from_single_column(sheet) -> Dict[str, object]:
    """Recover settings where key/value were incorrectly written into one column."""
    pairs: Dict[str, object] = {}
    if sheet.max_row < 2:
        return pairs
    for row_idx in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_idx, column=1).value
        if not cell_value:
            continue
        raw = str(cell_value).strip()
        match_key = None
        for key in _KNOWN_SETTINGS_COLUMNS:
            if raw.lower().startswith(key):
                match_key = key
                break
        if not match_key:
            continue
        value = raw[len(match_key):].strip()
        pairs[match_key] = value
    return pairs


def _normalize_settings_sheet(sheet) -> None:
    """Normalize settings sheet to row 1 headers and row 2 values."""
    header_row = _find_settings_header_row(sheet)
    if header_row:
        header_keys = {_normalize_header_name(v) for v in [cell.value for cell in sheet[header_row]] if v}
        if _KEY_VALUE_HEADERS.issubset(header_keys):
            pairs = _extract_pairs_from_key_value(sheet, header_row)
            if pairs:
                headers = list(pairs.keys())
                values = pairs.copy()
                _rebuild_settings_sheet(sheet, headers, values)
            return

        # Standard header row found (row 1 or later)
        headers_with_cols = []
        values: Dict[str, object] = {}
        for cell in sheet[header_row]:
            if cell.value:
                header_name = _normalize_header_name(cell.value)
                headers_with_cols.append((cell.column, header_name))
        if headers_with_cols:
            value_row = header_row + 1
            for col_idx, header in headers_with_cols:
                if value_row <= sheet.max_row:
                    cell_value = sheet.cell(row=value_row, column=col_idx).value
                    if header in PROTECTED_SETTINGS_COLUMNS and isinstance(cell_value, str) and cell_value.startswith("="):
                        values[header] = cell_value
                    else:
                        values[header] = cell_value
            headers = [header for _, header in headers_with_cols]
            _rebuild_settings_sheet(sheet, headers, values)
        return

    # Handle corrupted single-column key/value layout (e.g., column_namevalue)
    cell_a1 = sheet.cell(row=1, column=1).value
    if cell_a1:
        header_text = str(cell_a1).strip().lower().replace(" ", "")
        if "column_name" in header_text and "value" in header_text:
            pairs = _extract_pairs_from_single_column(sheet)
            if pairs:
                headers = list(pairs.keys())
                values = pairs.copy()
                _rebuild_settings_sheet(sheet, headers, values)
    return


def read_form_settings(xlsx_path: Path) -> Dict[str, str]:
    """Read form_title and form_id from settings sheet."""
    result = {"form_title": "", "form_id": ""}
    if not xlsx_path or not Path(xlsx_path).exists():
        return result

    wb = openpyxl.load_workbook(xlsx_path)
    sheet = _get_settings_sheet(wb)
    if sheet is None:
        return result

    header_map = _get_header_map(sheet)
    header_map_lower = {k.lower(): v for k, v in header_map.items()}

    if _KEY_VALUE_HEADERS.issubset(header_map_lower):
        name_col = header_map_lower["column_name"]
        value_col = header_map_lower["value"]
        for row_idx in range(2, sheet.max_row + 1):
            key_val = sheet.cell(row=row_idx, column=name_col).value
            value_val = sheet.cell(row=row_idx, column=value_col).value
            if key_val is None:
                continue
            key = str(key_val).strip().lower()
            if key in result and value_val is not None:
                result[key] = str(value_val).strip()
    else:
        row = 2
        for key in REQUIRED_SETTINGS_COLUMNS:
            col = header_map.get(key) or header_map_lower.get(key)
            if col:
                value = sheet.cell(row=row, column=col).value
                if value is not None:
                    result[key] = str(value).strip()
    return result


def ensure_settings_columns(sheet) -> Dict[str, int]:
    """Ensure required settings columns exist; return header map.

    Values are always written to row 2, aligned with row 1 headers.
    """
    _normalize_settings_sheet(sheet)
    header_map = _get_header_map(sheet)
    header_map_lower = {k.lower(): v for k, v in header_map.items()}
    if not header_map:
        # Create header row if missing
        for idx, col in enumerate(REQUIRED_SETTINGS_COLUMNS, start=1):
            sheet.cell(row=1, column=idx, value=col)
        return _get_header_map(sheet)

    next_col = len(header_map) + 1
    for col in REQUIRED_SETTINGS_COLUMNS:
        if col not in header_map and col not in header_map_lower:
            sheet.cell(row=1, column=next_col, value=col)
            header_map[col] = next_col
            next_col += 1

    return header_map


def _ensure_version_formula(sheet, header_map: Dict[str, int]) -> None:
    """Ensure version column exists and contains the required formula."""
    header_map_lower = {k.lower(): v for k, v in header_map.items()}
    version_col = header_map.get("version") or header_map_lower.get("version")
    if not version_col:
        version_col = len(header_map) + 1
        sheet.cell(row=1, column=version_col, value="version")
        header_map["version"] = version_col

    cell = sheet.cell(row=2, column=version_col)
    current = cell.value
    if current != VERSION_FORMULA:
        cell.value = VERSION_FORMULA


def set_form_settings(xlsx_path: Path, form_title: Optional[str] = None, form_id: Optional[str] = None) -> bool:
    """Create/update settings sheet and set form_title/form_id.

    Values are written to row 2, aligned with headers in row 1.
    """
    if not xlsx_path:
        return False

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        return False

    history_manager = None
    lock_acquired = False
    try:
        if HISTORY_AVAILABLE:
            history_manager = WorkbookHistoryManager(xlsform_path=xlsx_path, project_dir=xlsx_path.parent)
            history_manager.acquire_lock()
            lock_acquired = True
            history_manager.create_snapshot(
                action_type="update_settings",
                description="Pre-change snapshot before updating settings sheet",
                details=f"Target file: {xlsx_path.name}",
                command="python scripts/update_settings.py",
            )

        wb = openpyxl.load_workbook(xlsx_path)
        sheet = _get_settings_sheet(wb)
        if sheet is None:
            sheet = wb.create_sheet(SETTINGS_SHEET_NAME)

        header_map = ensure_settings_columns(sheet)
        _ensure_version_formula(sheet, header_map)
        header_map_lower = {k.lower(): v for k, v in header_map.items()}
        row = 2
        if form_title is not None:
            col = header_map.get("form_title") or header_map_lower.get("form_title")
            if col:
                sheet.cell(row=row, column=col, value=form_title)
        if form_id is not None:
            col = header_map.get("form_id") or header_map_lower.get("form_id")
            if col:
                sheet.cell(row=row, column=col, value=form_id)

        # Never overwrite protected columns like version (often formula-driven)
        for protected in PROTECTED_SETTINGS_COLUMNS:
            if protected in header_map:
                pass

        wb.save(xlsx_path)
        return True
    finally:
        if lock_acquired and history_manager is not None:
            history_manager.release_lock()


def missing_required_settings(xlsx_path: Path) -> List[str]:
    """Return list of missing required settings fields."""
    settings = read_form_settings(xlsx_path)
    missing = [key for key in REQUIRED_SETTINGS_COLUMNS if not settings.get(key)]
    return missing
