"""Add questions to XLSForm survey with best practices."""

import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl

# Ensure UTF-8 encoding for Windows console output
if sys.platform == 'win32':
    import locale
    import codecs
    # Try to set stdout/stderr encoding to UTF-8
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except AttributeError:
        # Python < 3.7 doesn't have reconfigure, use environment variable approach
        import os
        os.environ['PYTHONIOENCODING'] = 'utf-8'

# CRITICAL: Add scripts directory to Python path for sibling imports
# This allows the script to find sibling modules whether run from project root or scripts dir
_scripts_dir = Path(__file__).parent.resolve()
if str(_scripts_dir) not in sys.path:
    sys.path.insert(0, str(_scripts_dir))

# Try to import logger, fail gracefully if not available
try:
    from log_activity import ActivityLogger
    LOGGING_AVAILABLE = True
except ImportError:
    LOGGING_AVAILABLE = False

# Try to import config, fail gracefully if not available
try:
    from config import ProjectConfig
    CONFIG_AVAILABLE = True
except ImportError:
    CONFIG_AVAILABLE = False

# Try to import form structure, fail gracefully if not available
try:
    from form_structure import (
        find_insertion_point,
        freeze_top_row,
        find_header_row,
        build_column_mapping,
        ensure_blank_row_gap,
        is_metadata_field,
    )
    FORM_STRUCTURE_AVAILABLE = True
except ImportError:
    FORM_STRUCTURE_AVAILABLE = False

# Try to import display module, fail gracefully if not available
try:
    from display import print_questions_added
    DISPLAY_AVAILABLE = True
except ImportError:
    DISPLAY_AVAILABLE = False

# Try to import history manager, fail gracefully if not available
try:
    from history_manager import WorkbookHistoryManager
    HISTORY_AVAILABLE = True
except ImportError:
    HISTORY_AVAILABLE = False

# Try to import settings helpers, fail gracefully if not available
try:
    from settings_utils import ensure_settings_columns, VERSION_FORMULA
    SETTINGS_UTILS_AVAILABLE = True
except ImportError:
    SETTINGS_UTILS_AVAILABLE = False
    ensure_settings_columns = None
    VERSION_FORMULA = '=TEXT(NOW(), "yyyymmddhhmmss")'


# Try to import AI components, fail gracefully if not available
try:
    from knowledge_base.rag_engine import RAGEngine
    from question_type_analyzer import QuestionTypeAnalyzer
    from constraint_generator import SmartConstraintGenerator, Question
    from choice_optimizer import ChoiceListOptimizer
    from other_specify_handler import OtherSpecifyHandler, Question as OSQuestion
    AI_COMPONENTS_AVAILABLE = True
except ImportError:
    AI_COMPONENTS_AVAILABLE = False
    RAGEngine = None
    QuestionTypeAnalyzer = None
    SmartConstraintGenerator = None
    Question = None
    ChoiceListOptimizer = None
    OtherSpecifyHandler = None
    OSQuestion = None


def get_best_practices(question_type, question_name, question_label=""):
    """Apply best practices based on question type and name.

    Uses AI-powered SmartConstraintGenerator if available, otherwise
    falls back to rule-based logic.

    Args:
        question_type: XLSForm question type
        question_name: Question name/variable
        question_label: Question label text (for AI analysis)

    Returns:
        dict with constraint, constraint_message, required, required_message, appearance
    """
    # Use AI-powered constraint generator if available
    if AI_COMPONENTS_AVAILABLE and SmartConstraintGenerator:
        try:
            generator = SmartConstraintGenerator()
            question_obj = Question(question_type, question_name, question_label)
            constraints = generator.generate_constraints(question_obj)

            return {
                "constraint": constraints.constraint,
                "constraint_message": constraints.constraint_message,
                "required": constraints.required,
                "required_message": constraints.required_message,
                "appearance": constraints.appearance
            }
        except Exception as e:
            # Fall back to simple rules on error
            pass

    # Fallback to simple rule-based logic
    # Field types that should NOT be required (computed/read-only fields)
    non_input_types = {
        'calculate', 'hidden', 'note', 'imei', 'deviceid', 'subscriberid',
        'simserial', 'phonenumber', 'username', 'start', 'end', 'today',
        'audit', 'barcode', 'qrcode', 'image', 'audio', 'video', 'file'
    }

    is_non_input = question_type.lower() in non_input_types

    result = {
        "constraint": "",
        "constraint_message": "",
        "required": "" if is_non_input else "yes",
        "required_message": "" if is_non_input else "This field is required",
        "appearance": ""
    }

    # Name fields: no numbers, special chars
    if any(keyword in question_name.lower() for keyword in ["name", "first", "last", "full"]):
        result.update({
            "constraint": "regex(., '^[a-zA-Z\\\\s\\\\-\\\\\\\\.']$')",
            "constraint_message": "Please enter a valid name (letters only)",
            "required": "" if is_non_input else "yes",
            "required_message": "" if is_non_input else "Name is required"
        })

    # Age fields: reasonable range
    elif question_type == "integer" and "age" in question_name.lower():
        result.update({
            "constraint": ". >= 0 and . <= 130",
            "constraint_message": "Age must be between 0 and 130",
            "required": "" if is_non_input else "yes",
            "required_message": "" if is_non_input else "Age is required"
        })

    # Integer fields: non-negative by default
    elif question_type == "integer":
        result.update({
            "constraint": ". >= 0",
            "constraint_message": "Value must be zero or positive",
            "required": "" if is_non_input else "yes",
            "required_message": "" if is_non_input else "This field is required"
        })

    # Decimal fields: positive by default
    elif question_type == "decimal":
        result.update({
            "constraint": ". > 0",
            "constraint_message": "Value must be positive",
            "required": "" if is_non_input else "yes",
            "required_message": "" if is_non_input else "This field is required"
        })

    # Text fields: required by default (unless non-input type)
    elif question_type == "text":
        if not is_non_input:
            result.update({
                "required": "yes",
                "required_message": "This field is required"
            })

    # Select questions: always required (unless non-input type)
    elif question_type.startswith("select_"):
        if not is_non_input:
            result.update({
                "required": "yes",
                "required_message": "Please select an option"
            })

    return result


def _cell_has_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


_SELECT_TYPES = {"select_one", "select_multiple"}
_MAX_QUESTION_NAME_LEN = 32
_MAX_LIST_NAME_LEN = 24
_MAX_TOKEN_LEN = 10
_QUESTION_STOPWORDS = {
    "a", "an", "the", "is", "are", "was", "were", "be", "to", "for", "of", "and",
    "or", "in", "on", "at", "with", "from", "by", "as", "that", "this", "these",
    "those", "what", "which", "who", "whom", "whose", "when", "where", "why", "how",
    "do", "does", "did", "can", "could", "will", "would", "should", "your", "you",
    "please", "tell", "me",
}
_TOKEN_RE = re.compile(r"[a-z0-9]+")
_INVALID_IDENTIFIER_RE = re.compile(r"[^a-z0-9_]+")


def _to_ascii_identifier(
    raw_value: str,
    fallback: str = "question",
    max_length: int = 64,
    allow_trailing_numeric: bool = False,
) -> str:
    text = str(raw_value or "").strip().lower()
    text = text.replace("-", "_")
    text = _INVALID_IDENTIFIER_RE.sub("_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    if not text:
        text = fallback
    if text[0].isdigit():
        text = f"q_{text}"
    # Avoid trailing numeric suffixes in base variable/list names to reduce repeat/export ambiguity.
    if not allow_trailing_numeric and text[-1].isdigit():
        text = f"{text}_var"
    text = text[:max_length]
    if not allow_trailing_numeric and text and text[-1].isdigit():
        if len(text) >= max_length:
            text = text[:-1] + "v"
        else:
            text = f"{text}v"
    return text


def _suffix_letters(index: int) -> str:
    # 1 -> a, 2 -> b, ... 26 -> z, 27 -> aa ...
    if index <= 0:
        return "a"
    letters = []
    value = index
    while value > 0:
        value -= 1
        letters.append(chr(ord("a") + (value % 26)))
        value //= 26
    return "".join(reversed(letters))


def _ensure_unique_identifier(base_name: str, used_names: set, max_length: int = 64) -> str:
    candidate = _to_ascii_identifier(base_name, max_length=max_length)
    if candidate not in used_names:
        return candidate

    idx = 1
    while True:
        suffix = f"_{_suffix_letters(idx)}"
        base_trimmed = candidate[: max(1, max_length - len(suffix))]
        candidate_with_suffix = f"{base_trimmed}{suffix}"
        if candidate_with_suffix not in used_names:
            return candidate_with_suffix[:max_length]
        idx += 1


def _derive_name_from_label(label: str) -> str:
    text = str(label or "").strip().lower()
    text = re.sub(r"\$\{[^}]+\}", " ", text)
    tokens = [tok for tok in _TOKEN_RE.findall(text) if tok and tok not in _QUESTION_STOPWORDS]
    if not tokens:
        tokens = _TOKEN_RE.findall(text)
    if not tokens:
        return "question"

    compact_tokens = [tok[:_MAX_TOKEN_LEN] for tok in tokens[:4] if tok]
    if not compact_tokens:
        compact_tokens = tokens[:2]
    base = "_".join(compact_tokens)
    return _to_ascii_identifier(base, fallback="question", max_length=_MAX_QUESTION_NAME_LEN)


def _derive_short_list_name(question_name: str) -> str:
    base = _to_ascii_identifier(
        question_name,
        fallback="choices",
        max_length=max(_MAX_LIST_NAME_LEN, 16),
    )
    for suffix in ("_field", "_var", "_question", "_item"):
        if base.endswith(suffix):
            base = base[: -len(suffix)]
            break

    tokens = [tok for tok in base.split("_") if tok]
    if len(tokens) >= 3:
        compact = "_".join(tokens[:3])
    elif len(tokens) >= 2:
        compact = "_".join(tokens[:2])
    else:
        compact = base

    return _to_ascii_identifier(
        f"{compact}_opts",
        fallback="choices_opts",
        max_length=_MAX_LIST_NAME_LEN,
    )


def _split_select_type(raw_type: str) -> Tuple[str, str]:
    text = str(raw_type or "").strip().lower().replace("-", "_")
    if not text:
        return "", ""

    parts = [part for part in text.split() if part]
    if not parts:
        return "", ""

    base_type = parts[0]
    if base_type == "selectone":
        base_type = "select_one"
    elif base_type == "selectmultiple":
        base_type = "select_multiple"

    list_name = parts[1] if len(parts) > 1 else ""
    return base_type, list_name


def _normalize_choice_name(raw_value: str, label_value: str) -> str:
    if _cell_has_value(raw_value):
        text = str(raw_value).strip().lower().replace(" ", "_")
    else:
        text = str(label_value or "").strip().lower().replace(" ", "_")
    text = re.sub(r"[^a-z0-9_.-]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    if not text:
        text = "option"
    return text[:64]


def _normalize_choice_entries(raw_choices) -> List[Dict[str, str]]:
    if raw_choices is None:
        return []

    if isinstance(raw_choices, dict):
        # Not expected, but keep this permissive for parser variants.
        raw_choices = [raw_choices]

    if not isinstance(raw_choices, list):
        return []

    normalized: List[Dict[str, str]] = []
    seen_names: set = set()
    for entry in raw_choices:
        if isinstance(entry, dict):
            label = str(entry.get("label") or entry.get("name") or entry.get("value") or "").strip()
            value = entry.get("value")
            if value is None:
                value = entry.get("name")
            media_image = entry.get("media::image") or entry.get("image")
            media_audio = entry.get("media::audio") or entry.get("audio")
            media_video = entry.get("media::video") or entry.get("video")
        else:
            label = str(entry or "").strip()
            value = None
            media_image = None
            media_audio = None
            media_video = None

        if not label:
            continue

        choice_name = _normalize_choice_name(value, label)
        if choice_name in seen_names:
            continue
        seen_names.add(choice_name)

        choice_item: Dict[str, str] = {
            "name": choice_name,
            "label": label,
        }
        if _cell_has_value(media_image):
            choice_item["media::image"] = str(media_image).strip()
        if _cell_has_value(media_audio):
            choice_item["media::audio"] = str(media_audio).strip()
        if _cell_has_value(media_video):
            choice_item["media::video"] = str(media_video).strip()
        normalized.append(choice_item)

    return normalized


def _extract_questions_payload(data) -> List[Dict[str, object]]:
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    if isinstance(data, dict):
        questions = data.get("questions")
        if isinstance(questions, list):
            return [item for item in questions if isinstance(item, dict)]
        if isinstance(data.get("text"), str) or isinstance(data.get("label"), str):
            return [data]
    return []


def add_questions(
    questions_data,
    survey_file=None,
    name_strategy: str = "preserve",
):
    """Add questions to XLSForm with best practices.

    Args:
        questions_data: List/dict payload of questions. Supports parser JSON with `text`+`choices`.
        survey_file: Path to survey XLSForm file (optional, uses config if not specified)
        name_strategy: `preserve` or `semantic` for incoming question names.

    Returns:
        dict with success status and details
    """
    # Initialize config once for use throughout the function
    # This ensures consistency and supports custom project directories
    config = None
    if CONFIG_AVAILABLE:
        config = ProjectConfig()

    if name_strategy not in {"preserve", "semantic"}:
        return {"success": False, "error": "name_strategy must be 'preserve' or 'semantic'"}

    parsed_questions = _extract_questions_payload(questions_data)
    if not parsed_questions:
        return {
            "success": False,
            "error": "No question objects found. Provide a list or {'questions': [...]} payload.",
        }
    questions_data = parsed_questions

    history_manager = None
    lock_acquired = False
    snapshot_revision = ""

    try:
        # Determine file to use
        if survey_file is None:
            if config:
                survey_file = config.get_full_xlsform_path()
            else:
                survey_file = Path("survey.xlsx")
        else:
            survey_file = Path(survey_file)

        # Create immutable pre-change snapshot and acquire edit lock
        if HISTORY_AVAILABLE:
            project_dir = survey_file.parent
            if config:
                try:
                    if survey_file.resolve().is_relative_to(config.project_dir.resolve()):
                        project_dir = config.project_dir
                except Exception:
                    project_dir = survey_file.parent
            history_manager = WorkbookHistoryManager(
                xlsform_path=survey_file,
                project_dir=project_dir,
            )
            history_manager.acquire_lock()
            lock_acquired = True
            snapshot = history_manager.create_snapshot(
                action_type="add_questions",
                description=f"Pre-change snapshot before adding {len(questions_data)} question(s)",
                details=f"Target file: {survey_file.name}",
                command="/xlsform-add",
            )
            snapshot_revision = snapshot.get("revision_id", "")

        # Load workbook
        wb = openpyxl.load_workbook(survey_file)

        # Ensure settings.version uses formula default unless user explicitly overrides via update-settings.
        if SETTINGS_UTILS_AVAILABLE:
            try:
                if "settings" in wb.sheetnames:
                    settings_sheet = wb["settings"]
                else:
                    settings_sheet = wb.create_sheet("settings")
                settings_header_map = ensure_settings_columns(settings_sheet)
                settings_header_lower = {k.lower(): v for k, v in settings_header_map.items()}
                version_col = settings_header_map.get("version") or settings_header_lower.get("version")
                if version_col:
                    version_cell = settings_sheet.cell(row=2, column=version_col)
                    if version_cell.value is None or str(version_cell.value).strip() == "":
                        version_cell.value = VERSION_FORMULA
            except Exception:
                pass

        # Get survey sheet
        if "survey" not in wb.sheetnames:
            return {"success": False, "error": "'survey' sheet not found"}

        ws = wb["survey"]

        # Find header row
        if FORM_STRUCTURE_AVAILABLE:
            header_row = find_header_row(ws)
        else:
            header_row = None
            for row_idx in range(1, min(10, ws.max_row + 1)):
                cell_value = ws.cell(row_idx, 1).value
                if cell_value and str(cell_value).strip().lower() == "type":
                    header_row = row_idx
                    break

        if header_row is None:
            return {"success": False, "error": "Could not find header row with 'type' column"}

        # Build dynamic column mapping from actual header row
        # This maps column names to their actual positions (e.g., {"type": 1, "name": 3})
        if FORM_STRUCTURE_AVAILABLE:
            column_map = build_column_mapping(ws, header_row)
        else:
            # Fallback: assume standard column positions
            column_map = {
                "type": 1, "name": 2, "label": 3, "hint": 4, "required": 5,
                "calculation": 6, "relevant": 7, "constraint": 8,
                "constraint_message": 9, "required_message": 10, "appearance": 11,
                "default": 12, "media::image": 13, "media::audio": 14, "media::video": 15
            }

        # Validate required columns exist
        required_columns = ["type", "name", "label"]
        missing_columns = [col for col in required_columns if col not in column_map]
        if missing_columns:
            return {"success": False, "error": f"Missing required columns in header row: {', '.join(missing_columns)}"}

        def canonical_header_key(raw_key):
            key = str(raw_key or "").strip().lower()
            if not key:
                return ""
            media_aliases = {
                "image": "media::image",
                "media:image": "media::image",
                "media:_image": "media::image",
                "audio": "media::audio",
                "media:audio": "media::audio",
                "media:_audio": "media::audio",
                "video": "media::video",
                "media:video": "media::video",
                "media:_video": "media::video",
            }
            return media_aliases.get(key, key)

        # Build existing question-name set from current survey content.
        name_col = column_map.get("name", 2)
        existing_question_names = set()
        for row_idx in range(header_row + 1, ws.max_row + 1):
            name_val = ws.cell(row_idx, name_col).value
            if _cell_has_value(name_val):
                existing_question_names.add(str(name_val).strip())

        # Build existing choices-sheet state for list reuse/merge.
        existing_choice_lists: Dict[str, Dict[str, str]] = {}
        existing_choice_signatures: Dict[Tuple[Tuple[str, str], ...], str] = {}
        used_list_names: set = set()
        if "choices" in wb.sheetnames:
            ws_existing_choices = wb["choices"]
            choice_headers: Dict[str, int] = {}
            for col_idx in range(1, ws_existing_choices.max_column + 1):
                header_val = ws_existing_choices.cell(1, col_idx).value
                if not _cell_has_value(header_val):
                    continue
                normalized_header = canonical_header_key(str(header_val))
                if normalized_header:
                    choice_headers[normalized_header] = col_idx

            list_col = choice_headers.get("list_name")
            choice_name_col = choice_headers.get("name")
            choice_label_col = choice_headers.get("label")
            if list_col and choice_name_col and choice_label_col:
                for row_idx in range(2, ws_existing_choices.max_row + 1):
                    list_val = ws_existing_choices.cell(row_idx, list_col).value
                    name_val = ws_existing_choices.cell(row_idx, choice_name_col).value
                    label_val = ws_existing_choices.cell(row_idx, choice_label_col).value
                    if not (_cell_has_value(list_val) and _cell_has_value(name_val)):
                        continue
                    list_name_raw = str(list_val).strip()
                    choice_name_raw = _normalize_choice_name(name_val, label_val)
                    choice_label_raw = str(label_val or "").strip()
                    existing_choice_lists.setdefault(list_name_raw, {})[choice_name_raw] = choice_label_raw

                for list_name_raw, options_map in existing_choice_lists.items():
                    used_list_names.add(list_name_raw)
                    signature = tuple(sorted((k, v.lower()) for k, v in options_map.items()))
                    existing_choice_signatures[signature] = list_name_raw

        prepared_questions: List[Dict[str, object]] = []
        choice_batches: Dict[str, List[Dict[str, str]]] = {}
        used_question_names = set(existing_question_names)

        for raw_question in questions_data:
            if not isinstance(raw_question, dict):
                continue
            question = dict(raw_question)
            label_value = question.get("label")
            if not _cell_has_value(label_value):
                label_value = question.get("text")
            if not _cell_has_value(label_value):
                continue
            label_text = str(label_value).strip()
            question["label"] = label_text

            raw_name = question.get("name")
            if name_strategy == "semantic" or not _cell_has_value(raw_name):
                base_name = _derive_name_from_label(label_text)
            else:
                base_name = _to_ascii_identifier(
                    str(raw_name),
                    fallback=_derive_name_from_label(label_text),
                    max_length=_MAX_QUESTION_NAME_LEN,
                )
            unique_name = _ensure_unique_identifier(
                base_name,
                used_question_names,
                max_length=_MAX_QUESTION_NAME_LEN,
            )
            used_question_names.add(unique_name)
            question["name"] = unique_name

            normalized_choices = _normalize_choice_entries(question.get("choices"))
            base_type, inline_list_name = _split_select_type(question.get("type", ""))
            explicit_list_name = str(question.get("list_name") or inline_list_name or "").strip()

            if normalized_choices:
                if base_type not in _SELECT_TYPES:
                    base_type = "select_one"
                signature = tuple(sorted((c["name"], c["label"].lower()) for c in normalized_choices))

                if explicit_list_name:
                    if explicit_list_name in existing_choice_lists:
                        list_name = explicit_list_name
                    else:
                        list_name = _to_ascii_identifier(
                            explicit_list_name,
                            fallback=_derive_short_list_name(unique_name),
                            max_length=_MAX_LIST_NAME_LEN,
                        )
                else:
                    proposed = _derive_short_list_name(unique_name)
                    list_name = existing_choice_signatures.get(signature, proposed)
                    if list_name in used_list_names and list_name not in existing_choice_lists:
                        list_name = _ensure_unique_identifier(
                            list_name,
                            used_list_names,
                            max_length=_MAX_LIST_NAME_LEN,
                        )
                    if list_name in existing_choice_lists:
                        existing_signature = tuple(
                            sorted((k, v.lower()) for k, v in existing_choice_lists[list_name].items())
                        )
                        if existing_signature != signature:
                            list_name = _ensure_unique_identifier(
                                list_name,
                                used_list_names,
                                max_length=_MAX_LIST_NAME_LEN,
                            )

                used_list_names.add(list_name)
                question["type"] = f"{base_type} {list_name}"
                existing_choice_signatures.setdefault(signature, list_name)

                bucket = choice_batches.setdefault(list_name, [])
                bucket_names = {item["name"] for item in bucket}
                existing_names_for_list = set(existing_choice_lists.get(list_name, {}).keys())
                for choice_item in normalized_choices:
                    if choice_item["name"] in bucket_names or choice_item["name"] in existing_names_for_list:
                        continue
                    bucket.append(choice_item)
                    bucket_names.add(choice_item["name"])
                    existing_choice_lists.setdefault(list_name, {})[choice_item["name"]] = choice_item["label"]
            else:
                if base_type in _SELECT_TYPES:
                    if explicit_list_name in existing_choice_lists:
                        list_name = explicit_list_name
                    else:
                        list_name = _to_ascii_identifier(
                            explicit_list_name,
                            fallback="choices_opts",
                            max_length=_MAX_LIST_NAME_LEN,
                        )
                    if list_name:
                        question["type"] = f"{base_type} {list_name}"
                    else:
                        question["type"] = "text"
                else:
                    question["type"] = base_type or "text"

            prepared_questions.append(question)

        if not prepared_questions:
            return {
                "success": True,
                "added": [],
                "total": 0,
                "message": "No valid questions found to add",
            }

        questions_data = prepared_questions

        # Add optional headers when incoming payload contains columns not yet present.
        # This is especially important for imported media fields (media::image/audio/video).
        skip_dynamic_fields = {
            "type", "name", "label", "constraint", "constraint_message",
            "required", "required_message", "choices", "number", "page", "text", "list_name",
        }
        dynamic_headers = []
        for q in questions_data:
            for key in q.keys():
                canonical_key = canonical_header_key(key)
                if not canonical_key or canonical_key in skip_dynamic_fields:
                    continue
                if canonical_key not in column_map and canonical_key not in dynamic_headers:
                    dynamic_headers.append(canonical_key)

        if dynamic_headers:
            next_col = max(column_map.values()) + 1 if column_map else 1
            for header_name in dynamic_headers:
                ws.cell(header_row, next_col, header_name)
                column_map[header_name] = next_col
                next_col += 1

        # Find insertion point using smart logic (or fallback to simple logic)
        if FORM_STRUCTURE_AVAILABLE:
            insertion_row = find_insertion_point(ws, header_row, questions_data, column_map)
            has_metadata = any(is_metadata_field(q.get("type", "")) for q in questions_data)
            if not has_metadata:
                check_columns = [column_map.get("type", 1), column_map.get("name", 2)]
                insertion_row = ensure_blank_row_gap(ws, insertion_row, header_row, check_columns, min_blank_rows=1)
        else:
            # Fallback: find last data row
            name_col = column_map.get("name", 2)  # Use dynamic mapping or fallback to column 2
            insertion_row = header_row + 1
            for row_idx in range(header_row + 1, ws.max_row + 1):
                name_val = ws.cell(row_idx, name_col).value
                if name_val is not None and str(name_val).strip():
                    insertion_row = row_idx + 1

        # Add questions
        added = []
        current_row = insertion_row
        for q in questions_data:
            q_type = q.get("type", "text")
            q_name = q.get("name", "")
            q_label = q.get("label", "")

            # Apply best practices (with AI enhancement if available)
            practices = get_best_practices(q_type, q_name, q_label)

            # Allow override if user specified constraints/required
            constraint = q.get("constraint", practices["constraint"])
            constraint_msg = q.get("constraint_message", practices["constraint_message"])
            required = q.get("required", practices["required"])
            required_msg = q.get("required_message", practices["required_message"])
            appearance = q.get("appearance", practices.get("appearance", ""))

            # Set core values (required columns always exist due to validation above)
            ws.cell(current_row, column_map["type"], q_type)
            ws.cell(current_row, column_map["name"], q_name)
            ws.cell(current_row, column_map["label"], q_label)

            # Set constraint fields (optional - check if column exists)
            if constraint and "constraint" in column_map:
                ws.cell(current_row, column_map["constraint"], constraint)
            if constraint_msg and "constraint_message" in column_map:
                ws.cell(current_row, column_map["constraint_message"], constraint_msg)
            if required and "required" in column_map:
                ws.cell(current_row, column_map["required"], required)
            if required_msg and "required_message" in column_map:
                ws.cell(current_row, column_map["required_message"], required_msg)
            # Set appearance from best practices if not overridden by user
            if appearance and "appearance" not in q and "appearance" in column_map:
                ws.cell(current_row, column_map["appearance"], appearance)

            # Handle additional fields dynamically from question data
            # This supports: hint, calculation, relevant, appearance, default, media::image, media::audio, media::video, etc.
            for key, value in q.items():
                if key in ["type", "name", "label", "constraint", "constraint_message", "required", "required_message"]:
                    continue
                if key in ["choices", "number", "page", "text", "list_name"]:
                    continue

                key_lower = canonical_header_key(key)
                if key_lower in column_map:
                    ws.cell(current_row, column_map[key_lower], value)

            added.append({
                "row": current_row,
                "type": q_type,
                "name": q_name,
                "label": q_label,
                "required": required,
                "constraint": constraint
            })

            current_row += 1

        choices_added: List[Dict[str, object]] = []
        if choice_batches:
            if "choices" in wb.sheetnames:
                ws_choices = wb["choices"]
            else:
                ws_choices = wb.create_sheet("choices")
                ws_choices.cell(1, 1, "list_name")
                ws_choices.cell(1, 2, "name")
                ws_choices.cell(1, 3, "label")

            choices_column_map: Dict[str, int] = {}
            for col_idx in range(1, ws_choices.max_column + 1):
                header_value = ws_choices.cell(1, col_idx).value
                if not _cell_has_value(header_value):
                    continue
                normalized_header = canonical_header_key(str(header_value))
                if normalized_header:
                    choices_column_map[normalized_header] = col_idx

            required_choice_headers = ["list_name", "name", "label"]
            next_choice_col = max(choices_column_map.values()) + 1 if choices_column_map else 1
            for header_name in required_choice_headers:
                if header_name in choices_column_map:
                    continue
                ws_choices.cell(1, next_choice_col, header_name)
                choices_column_map[header_name] = next_choice_col
                next_choice_col += 1

            # Add media columns only when needed.
            needed_choice_headers = set()
            for choice_items in choice_batches.values():
                for choice_item in choice_items:
                    if _cell_has_value(choice_item.get("media::image")):
                        needed_choice_headers.add("media::image")
                    if _cell_has_value(choice_item.get("media::audio")):
                        needed_choice_headers.add("media::audio")
                    if _cell_has_value(choice_item.get("media::video")):
                        needed_choice_headers.add("media::video")

            for header_name in sorted(needed_choice_headers):
                if header_name in choices_column_map:
                    continue
                ws_choices.cell(1, next_choice_col, header_name)
                choices_column_map[header_name] = next_choice_col
                next_choice_col += 1

            # Build up-to-date lookup (existing + newly appended) to avoid duplicates.
            existing_by_list: Dict[str, Dict[str, int]] = {}
            list_col = choices_column_map["list_name"]
            name_col_choices = choices_column_map["name"]
            for row_idx in range(2, ws_choices.max_row + 1):
                list_val = ws_choices.cell(row_idx, list_col).value
                name_val = ws_choices.cell(row_idx, name_col_choices).value
                if not (_cell_has_value(list_val) and _cell_has_value(name_val)):
                    continue
                list_name = str(list_val).strip()
                choice_name = _normalize_choice_name(name_val, "")
                existing_by_list.setdefault(list_name, {})[choice_name] = row_idx

            append_row = ws_choices.max_row + 1
            for list_name, choice_items in choice_batches.items():
                for choice_item in choice_items:
                    choice_name = _normalize_choice_name(choice_item.get("name"), choice_item.get("label"))
                    if choice_name in existing_by_list.get(list_name, {}):
                        continue

                    ws_choices.cell(append_row, choices_column_map["list_name"], list_name)
                    ws_choices.cell(append_row, choices_column_map["name"], choice_name)
                    ws_choices.cell(append_row, choices_column_map["label"], choice_item.get("label", ""))

                    if "media::image" in choices_column_map and _cell_has_value(choice_item.get("media::image")):
                        ws_choices.cell(
                            append_row,
                            choices_column_map["media::image"],
                            str(choice_item.get("media::image")).strip(),
                        )
                    if "media::audio" in choices_column_map and _cell_has_value(choice_item.get("media::audio")):
                        ws_choices.cell(
                            append_row,
                            choices_column_map["media::audio"],
                            str(choice_item.get("media::audio")).strip(),
                        )
                    if "media::video" in choices_column_map and _cell_has_value(choice_item.get("media::video")):
                        ws_choices.cell(
                            append_row,
                            choices_column_map["media::video"],
                            str(choice_item.get("media::video")).strip(),
                        )

                    existing_by_list.setdefault(list_name, {})[choice_name] = append_row
                    choices_added.append(
                        {"row": append_row, "list_name": list_name, "name": choice_name, "label": choice_item.get("label", "")}
                    )
                    append_row += 1

        # Freeze the header row for better usability
        if FORM_STRUCTURE_AVAILABLE:
            try:
                freeze_top_row(ws, header_row)
            except Exception as e:
                # Fail gracefully if freeze panes doesn't work
                pass

        # Save workbook
        wb.save(survey_file)

        # Log activity
        if LOGGING_AVAILABLE and config:
            try:
                # Check if activity logging is enabled in config
                if config.is_activity_logging_enabled():
                    # Use project directory from config
                    # This respects the project_dir setting from xlsform-ai.json
                    # and supports both relative and absolute paths
                    logger = ActivityLogger(project_dir=config.project_dir)
                    questions_summary = ", ".join([f"{q['name']} ({q['type']})" for q in added])
                    detail_lines = [
                        f"Questions: {questions_summary}",
                        f"Rows: {', '.join([str(q['row']) for q in added])}",
                    ]
                    if choices_added:
                        detail_lines.append(
                            f"Choices added: {len(choices_added)} across {len({item['list_name'] for item in choices_added})} list(s)"
                        )
                    if snapshot_revision:
                        detail_lines.append(f"Snapshot revision: {snapshot_revision}")
                    logger.log_action(
                        action_type="add_questions",
                        description=f"Added {len(added)} question(s)",
                        details="\n".join(detail_lines),
                    )
                    log_file = logger.log_file
                    print(f"\n[OK] Activity logged to: {log_file.name}")
            except Exception as e:
                print(f"\n[WARNING] Could not log activity: {e}")


        return {
            "success": True,
            "added": added,
            "total": len(added),
            "choices_added": choices_added,
            "choices_total": len(choices_added),
            "snapshot_revision": snapshot_revision,
        }

    except Exception as e:
        return {"success": False, "error": str(e)}
    finally:
        if lock_acquired and history_manager is not None:
            history_manager.release_lock()


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='Add questions to XLSForm')
    parser.add_argument(
        'questions',
        nargs='?',
        help='JSON string of questions to add (or parser output payload)',
    )
    parser.add_argument(
        '--from-json-file',
        help='Read questions payload from JSON file (supports {"questions":[...]} and list payloads)',
    )
    parser.add_argument(
        '--name-strategy',
        choices=['preserve', 'semantic'],
        default='preserve',
        help='Question naming strategy for incoming payloads (default: preserve)',
    )
    parser.add_argument('--file', '-f',
                       help='Override XLSForm file name (default: use config or survey.xlsx)')

    args = parser.parse_args()

    # Parse question data from command line argument
    try:
        payload = None
        if args.from_json_file:
            payload_path = Path(args.from_json_file).resolve()
            if not payload_path.exists():
                print(f"ERROR: JSON file not found: {payload_path}")
                sys.exit(1)
            payload = json.loads(payload_path.read_text(encoding='utf-8-sig'))
        elif args.questions:
            payload = json.loads(args.questions)
        else:
            print("ERROR: Provide question JSON as positional argument or use --from-json-file")
            sys.exit(1)

        result = add_questions(
            payload,
            survey_file=args.file,
            name_strategy=args.name_strategy,
        )

        if result["success"]:
            # Use beautiful display if available
            if DISPLAY_AVAILABLE:
                print_questions_added(result['total'], result['added'])
            else:
                # Fallback to simple text output
                print(f"SUCCESS: Added {result['total']} question(s)")
                for q in result["added"]:
                    print(f"  Row {q['row']}: {q['type']} | {q['name']} | \"{q['label']}\"")
                    if q['required']:
                        print(f"    Required: {q['required']}")
                    if q['constraint']:
                        print(f"    Constraint: {q['constraint']}")
                if result.get("choices_total", 0):
                    print(f"  Choices added: {result['choices_total']}")

        else:
            print(f"ERROR: {result['error']}")
            sys.exit(1)

    except json.JSONDecodeError:
        print("ERROR: Invalid JSON format for question data")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: {str(e)}")
        sys.exit(1)
