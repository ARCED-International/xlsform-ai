#!/usr/bin/env python3
"""
XLSForm translation workflow for multilingual forms.

Core features:
1) Parse natural language intents (for /xlsform-translate style commands)
2) Add/normalize translation headers in row 1 without shifting existing values
3) Translate survey/choices text columns with strict row-1 column mapping
4) Update settings.default_language safely
5) Emit structured REPL output for agents
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except AttributeError:
        pass

_SCRIPTS_DIR = Path(__file__).parent.resolve()
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

try:
    from form_structure import normalize_header_name
except Exception:
    def normalize_header_name(value: str) -> str:
        return str(value).strip().lower().replace(" ", "_")

try:
    from config import ProjectConfig
    CONFIG_AVAILABLE = True
except Exception:
    CONFIG_AVAILABLE = False

try:
    from log_activity import ActivityLogger
    LOGGING_AVAILABLE = True
except Exception:
    LOGGING_AVAILABLE = False

try:
    from history_manager import WorkbookHistoryManager
    HISTORY_AVAILABLE = True
except Exception:
    HISTORY_AVAILABLE = False


TRANSLATABLE_CATEGORY_BASES: Dict[str, List[str]] = {
    "label": ["label"],
    "hint": ["hint"],
    "guidance_hint": ["guidance_hint"],
    "constraint_message": ["constraint_message"],
    "required_message": ["required_message"],
    "image": ["image", "media::image"],
    "audio": ["audio", "media::audio"],
    "video": ["video", "media::video"],
}

CATEGORY_ORDER = list(TRANSLATABLE_CATEGORY_BASES.keys())
TEXT_CATEGORIES = {
    "label",
    "hint",
    "guidance_hint",
    "constraint_message",
    "required_message",
}
MEDIA_CATEGORIES = {"image", "audio", "video"}

BASE_TO_CATEGORY: Dict[str, str] = {}
for _cat, _bases in TRANSLATABLE_CATEGORY_BASES.items():
    for _base in _bases:
        BASE_TO_CATEGORY[_base] = _cat

PLACEHOLDER_PATTERN = re.compile(r"\$\{[^}]+\}")
LANG_CODE_PATTERN = re.compile(r"^[a-z]{2,3}(?:-[a-z0-9]{2,8})?$", re.IGNORECASE)


@dataclass
class LanguageSpec:
    code: str
    display_name: str

    @property
    def header_label(self) -> str:
        return f"{self.display_name} ({self.code})"


@dataclass
class HeaderEntry:
    col: int
    raw: str
    base_raw: str
    base_normalized: str
    category: Optional[str]
    is_translation: bool
    language_code: Optional[str]
    language_display: Optional[str]


LANGUAGE_REGISTRY: Dict[str, Dict[str, List[str] | str]] = {
    "en": {"display": "English", "aliases": ["english", "eng", "en"]},
    "bn": {"display": "Bangla", "aliases": ["bangla", "bengali", "bn"]},
    "es": {"display": "Spanish", "aliases": ["spanish", "espanol", "es"]},
    "fr": {"display": "French", "aliases": ["french", "francais", "fr"]},
    "ar": {"display": "Arabic", "aliases": ["arabic", "ar"]},
    "pt": {"display": "Portuguese", "aliases": ["portuguese", "portugues", "pt"]},
    "sw": {"display": "Swahili", "aliases": ["swahili", "kiswahili", "sw"]},
    "hi": {"display": "Hindi", "aliases": ["hindi", "hi"]},
    "ur": {"display": "Urdu", "aliases": ["urdu", "ur"]},
    "ne": {"display": "Nepali", "aliases": ["nepali", "ne"]},
    "my": {"display": "Burmese", "aliases": ["burmese", "myanmar", "my"]},
}


def _normalize_alias(value: str) -> str:
    lowered = unicodedata.normalize("NFKD", str(value).strip().lower())
    stripped = "".join(ch for ch in lowered if not unicodedata.combining(ch))
    stripped = re.sub(r"\s+", " ", stripped)
    return stripped.strip()


LANGUAGE_ALIAS_INDEX: Dict[str, str] = {}
for _code, _meta in LANGUAGE_REGISTRY.items():
    LANGUAGE_ALIAS_INDEX[_normalize_alias(_code)] = _code
    LANGUAGE_ALIAS_INDEX[_normalize_alias(str(_meta["display"]))] = _code
    for _alias in _meta["aliases"]:
        LANGUAGE_ALIAS_INDEX[_normalize_alias(str(_alias))] = _code


def _cell_has_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _extract_parenthesized_code(token: str) -> Optional[str]:
    match = re.search(r"\(([^()]+)\)\s*$", token.strip())
    if not match:
        return None
    maybe_code = match.group(1).strip().lower()
    if LANG_CODE_PATTERN.match(maybe_code):
        return maybe_code
    return None


def _extract_loose_code(token: str) -> Optional[str]:
    token = token.strip().lower()
    if LANG_CODE_PATTERN.match(token):
        return token
    return None


def resolve_language_token(token: str, allow_unknown_code: bool = False) -> Tuple[Optional[LanguageSpec], Optional[str]]:
    raw = str(token or "").strip()
    if not raw:
        return None, "Empty language token."

    code = _extract_parenthesized_code(raw)
    display = raw
    if code:
        display = raw[: raw.rfind("(")].strip()
        if not display and code in LANGUAGE_REGISTRY:
            display = str(LANGUAGE_REGISTRY[code]["display"])

    normalized = _normalize_alias(display)
    if normalized in LANGUAGE_ALIAS_INDEX and not code:
        code = LANGUAGE_ALIAS_INDEX[normalized]

    if not code:
        loose = _extract_loose_code(raw)
        if loose:
            code = loose
            if loose in LANGUAGE_REGISTRY:
                display = str(LANGUAGE_REGISTRY[loose]["display"])
            else:
                display = loose

    if not code:
        return None, f"Could not resolve language code for '{raw}'. Use format like 'Spanish (es)'."

    if code not in LANGUAGE_REGISTRY and not allow_unknown_code:
        return None, f"Unsupported language code '{code}' for token '{raw}'."

    if code in LANGUAGE_REGISTRY:
        # Keep canonical ASCII-friendly display labels by default.
        display = str(LANGUAGE_REGISTRY[code]["display"])

    if not display:
        display = code

    return LanguageSpec(code=code, display_name=display), None


def _split_language_tokens(text: str) -> List[str]:
    if not text:
        return []
    normalized = re.sub(r"\band\b", ",", text, flags=re.IGNORECASE)
    tokens = [part.strip(" .") for part in normalized.split(",")]
    return [t for t in tokens if t]


def parse_instruction_intent(instruction: str) -> Dict[str, object]:
    cleaned = str(instruction or "").strip()
    if cleaned.lower().startswith("/xlsform-translate"):
        cleaned = cleaned[len("/xlsform-translate"):].strip()
    lowered = cleaned.lower()

    intent = "generic_translate"
    mode = "add-missing"
    language_tokens: List[str] = []

    add_match = re.search(r"\badd\s+(.+?)\s+language\b", cleaned, flags=re.IGNORECASE)
    if add_match:
        intent = "add_language"
        mode = "update-all"
        language_tokens = _split_language_tokens(add_match.group(1))
        return {"intent": intent, "mode": mode, "language_tokens": language_tokens, "instruction": cleaned}

    remaining_match = re.search(
        r"\b(?:do|complete|finish|translate)\s+(?:the\s+)?(.+?)\s+translations?\b",
        cleaned,
        flags=re.IGNORECASE,
    )
    if remaining_match and "remaining" in lowered:
        intent = "translate_remaining"
        mode = "add-missing"
        language_tokens = _split_language_tokens(remaining_match.group(1))
        return {"intent": intent, "mode": mode, "language_tokens": language_tokens, "instruction": cleaned}

    if "validate" in lowered and "translation" in lowered:
        intent = "validate_only"
        mode = "validate-only"

    if "header" in lowered and ("sync" in lowered or "normalize" in lowered):
        intent = "header_sync"
        mode = "header-sync"

    return {"intent": intent, "mode": mode, "language_tokens": language_tokens, "instruction": cleaned}


def _normalize_base_name(raw: str) -> str:
    return normalize_header_name(raw)


def _header_equals(left: str, right: str) -> bool:
    return _normalize_alias(left) == _normalize_alias(right)


def _parse_header_entry(raw_header: str, col: int) -> HeaderEntry:
    raw = str(raw_header).strip()
    segments = raw.split("::")

    if len(segments) >= 2:
        base_raw = "::".join(segments[:-1]).strip()
        lang_raw = segments[-1].strip()
        base_normalized = _normalize_base_name(base_raw)
        category = BASE_TO_CATEGORY.get(base_normalized)

        lang_spec, _ = resolve_language_token(lang_raw, allow_unknown_code=True)
        language_code = None
        language_display = None
        if lang_spec:
            language_code = lang_spec.code
            language_display = lang_spec.display_name
        else:
            language_code = _extract_parenthesized_code(lang_raw) or _extract_loose_code(lang_raw)
            language_display = lang_raw if language_code else None

        if category and language_code:
            return HeaderEntry(
                col=col,
                raw=raw,
                base_raw=base_raw,
                base_normalized=base_normalized,
                category=category,
                is_translation=True,
                language_code=language_code,
                language_display=language_display,
            )

    base_normalized = _normalize_base_name(raw)
    category = BASE_TO_CATEGORY.get(base_normalized)
    return HeaderEntry(
        col=col,
        raw=raw,
        base_raw=raw,
        base_normalized=base_normalized,
        category=category,
        is_translation=False,
        language_code=None,
        language_display=None,
    )


def _collect_headers(sheet) -> List[HeaderEntry]:
    entries: List[HeaderEntry] = []
    max_col = max(sheet.max_column, 1)
    for col in range(1, max_col + 1):
        value = sheet.cell(row=1, column=col).value
        if not _cell_has_value(value):
            continue
        entries.append(_parse_header_entry(str(value), col))
    return entries


def _count_non_empty_values(sheet, col: int) -> int:
    count = 0
    for row_idx in range(2, sheet.max_row + 1):
        if _cell_has_value(sheet.cell(row=row_idx, column=col).value):
            count += 1
    return count


def _choose_base_header(entries: List[HeaderEntry], category: str) -> str:
    aliases = TRANSLATABLE_CATEGORY_BASES[category]
    base_entries = [e for e in entries if (not e.is_translation and e.category == category)]

    for alias in aliases:
        for entry in base_entries:
            if entry.base_normalized == alias:
                return entry.base_raw

    trans_entries = [e for e in entries if (e.is_translation and e.category == category)]
    if trans_entries:
        return trans_entries[0].base_raw

    return aliases[0]


def _canonical_translation_header(base_header: str, language: LanguageSpec) -> str:
    return f"{base_header}::{language.display_name} ({language.code})"


def _merge_translation_columns(
    sheet,
    src_col: int,
    dst_col: int,
    sheet_name: str,
    category: str,
    dry_run: bool,
) -> Dict[str, object]:
    merged = 0
    conflicts: List[str] = []
    conflict_cap = 25

    for row_idx in range(2, sheet.max_row + 1):
        src_val = sheet.cell(row=row_idx, column=src_col).value
        dst_val = sheet.cell(row=row_idx, column=dst_col).value

        if not _cell_has_value(src_val):
            continue

        if not _cell_has_value(dst_val):
            if not dry_run:
                sheet.cell(row=row_idx, column=dst_col, value=src_val)
            merged += 1
            continue

        if str(src_val).strip() != str(dst_val).strip() and len(conflicts) < conflict_cap:
            conflicts.append(
                f"{sheet_name}: row {row_idx} has conflicting values for {category} across duplicate language columns."
            )

    return {"merged": merged, "conflicts": conflicts}


def ensure_language_columns(
    sheet,
    sheet_name: str,
    language: LanguageSpec,
    dry_run: bool,
) -> Dict[str, object]:
    entries = _collect_headers(sheet)
    max_col = max([entry.col for entry in entries], default=0)

    added_headers: List[str] = []
    renamed_headers: List[str] = []
    header_conflicts: List[str] = []
    merged_values = 0
    target_columns: Dict[str, int] = {}

    for category in CATEGORY_ORDER:
        base_header = _choose_base_header(entries, category)
        expected_header = _canonical_translation_header(base_header, language)

        candidates = [
            entry
            for entry in entries
            if entry.is_translation and entry.category == category and entry.language_code == language.code
        ]

        expected_entry = next((entry for entry in candidates if _header_equals(entry.raw, expected_header)), None)

        canonical_col: Optional[int] = None
        if expected_entry:
            canonical_col = expected_entry.col
        elif candidates:
            primary = candidates[0]
            canonical_col = primary.col

            expected_exists_elsewhere = any(_header_equals(entry.raw, expected_header) for entry in entries)
            if expected_exists_elsewhere and not _header_equals(primary.raw, expected_header):
                header_conflicts.append(
                    f"{sheet_name}: duplicate language headers detected for {category} ({language.code})."
                )
            elif not _header_equals(primary.raw, expected_header):
                if not dry_run:
                    sheet.cell(row=1, column=primary.col, value=expected_header)
                renamed_headers.append(f"{primary.raw} -> {expected_header}")
                updated = _parse_header_entry(expected_header, primary.col)
                entries = [updated if entry.col == primary.col else entry for entry in entries]
        else:
            max_col += 1
            canonical_col = max_col
            if not dry_run:
                sheet.cell(row=1, column=canonical_col, value=expected_header)
            added_headers.append(expected_header)
            entries.append(_parse_header_entry(expected_header, canonical_col))

        if canonical_col is None:
            continue

        target_columns[category] = canonical_col

        duplicate_candidates = [entry for entry in candidates if entry.col != canonical_col]
        for duplicate in duplicate_candidates:
            merge_result = _merge_translation_columns(
                sheet=sheet,
                src_col=duplicate.col,
                dst_col=canonical_col,
                sheet_name=sheet_name,
                category=category,
                dry_run=dry_run,
            )
            merged_values += int(merge_result["merged"])
            header_conflicts.extend(merge_result["conflicts"])

    return {
        "target_columns": target_columns,
        "added_headers": added_headers,
        "renamed_headers": renamed_headers,
        "header_conflicts": header_conflicts,
        "merged_values": merged_values,
    }


def _choose_source_column(
    sheet,
    entries: List[HeaderEntry],
    category: str,
    source_language: LanguageSpec,
) -> Optional[int]:
    translated_candidates = [
        entry
        for entry in entries
        if entry.is_translation and entry.category == category and entry.language_code == source_language.code
    ]
    if translated_candidates:
        ranked = sorted(
            translated_candidates,
            key=lambda entry: _count_non_empty_values(sheet, entry.col),
            reverse=True,
        )
        return ranked[0].col

    base_candidates = [entry for entry in entries if (not entry.is_translation and entry.category == category)]
    aliases = TRANSLATABLE_CATEGORY_BASES[category]
    for alias in aliases:
        for entry in base_candidates:
            if entry.base_normalized == alias:
                return entry.col
    if base_candidates:
        return base_candidates[0].col

    return None


def _build_translation_memory(
    sheet,
    source_columns: Dict[str, Optional[int]],
    target_columns: Dict[str, int],
) -> Dict[Tuple[str, str], str]:
    memory: Dict[Tuple[str, str], str] = {}
    for category in CATEGORY_ORDER:
        src_col = source_columns.get(category)
        dst_col = target_columns.get(category)
        if not src_col or not dst_col:
            continue
        for row_idx in range(2, sheet.max_row + 1):
            src_val = sheet.cell(row=row_idx, column=src_col).value
            dst_val = sheet.cell(row=row_idx, column=dst_col).value
            if not (_cell_has_value(src_val) and _cell_has_value(dst_val)):
                continue
            src_text = str(src_val).strip()
            dst_text = str(dst_val).strip()
            if src_text and dst_text:
                memory[(category, src_text)] = dst_text
    return memory


def _normalize_translation_lookup_text(text: str) -> str:
    normalized = str(text or "").strip()
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized


def _placeholder_tokens(text: str) -> Tuple[str, ...]:
    tokens = tuple(sorted(set(PLACEHOLDER_PATTERN.findall(str(text or "")))))
    return tokens


def _placeholders_match(source_text: str, translated_text: str) -> bool:
    return _placeholder_tokens(source_text) == _placeholder_tokens(translated_text)


def _prepare_translation_map(raw_payload: Optional[Dict[str, Any]]) -> Tuple[Dict[str, Dict[str, str]], List[str]]:
    warnings: List[str] = []
    prepared: Dict[str, Dict[str, str]] = {"*": {}}
    if not raw_payload:
        return prepared, warnings

    if not isinstance(raw_payload, dict):
        return prepared, ["Translation map must be a JSON object."]

    # Shortcut format:
    # { "How old are you?": "Bangla text", ... }
    if raw_payload and all(isinstance(value, str) for value in raw_payload.values()):
        for source_text, translated_text in raw_payload.items():
            src_key = _normalize_translation_lookup_text(str(source_text))
            dst_value = str(translated_text).strip()
            if src_key and dst_value:
                prepared["*"][src_key] = dst_value
        return prepared, warnings

    for key, value in raw_payload.items():
        normalized_key = _normalize_base_name(str(key))
        if normalized_key in {"global", "all", "any", "*"}:
            bucket = "*"
        elif normalized_key in CATEGORY_ORDER:
            bucket = normalized_key
        elif isinstance(value, str):
            src_key = _normalize_translation_lookup_text(str(key))
            dst_value = str(value).strip()
            if src_key and dst_value:
                prepared["*"][src_key] = dst_value
            continue
        else:
            warnings.append(f"Ignored unknown translation map key '{key}'.")
            continue

        if not isinstance(value, dict):
            warnings.append(f"Ignored translation map bucket '{key}' because it is not an object.")
            continue

        bucket_map = prepared.setdefault(bucket, {})
        for source_text, translated_text in value.items():
            if translated_text is None:
                continue
            src_key = _normalize_translation_lookup_text(str(source_text))
            dst_value = str(translated_text).strip()
            if src_key and dst_value:
                bucket_map[src_key] = dst_value

    return prepared, warnings


def _translation_map_entry_count(translation_map: Dict[str, Dict[str, str]]) -> int:
    return sum(len(bucket) for bucket in translation_map.values())


def _lookup_translation_map(
    translation_map: Dict[str, Dict[str, str]],
    category: str,
    source_text: str,
) -> Optional[str]:
    key = _normalize_translation_lookup_text(source_text)
    if not key:
        return None

    bucket = translation_map.get(category, {})
    if key in bucket:
        return bucket[key]

    global_bucket = translation_map.get("*", {})
    return global_bucket.get(key)


def _mask_placeholders(text: str) -> Tuple[str, Dict[str, str]]:
    tokens: Dict[str, str] = {}

    def replacer(match):
        key = f"__XLSFORM_TOKEN_{len(tokens)}__"
        tokens[key] = match.group(0)
        return key

    masked = PLACEHOLDER_PATTERN.sub(replacer, text)
    return masked, tokens


def _unmask_placeholders(text: str, tokens: Dict[str, str]) -> str:
    result = text
    for key, value in tokens.items():
        result = result.replace(key, value)
    return result


class GoogleTranslatorBackend:
    def __init__(self, source_code: str, target_code: str):
        try:
            from deep_translator import GoogleTranslator
        except Exception as exc:
            raise RuntimeError(
                "deep-translator is required for automatic translation. "
                "Install with: pip install -e \".[translate]\" (or pip install deep-translator)"
            ) from exc

        src = source_code if source_code else "auto"
        self._translator = GoogleTranslator(source=src, target=target_code)
        self._cache: Dict[str, str] = {}

    def translate_text(self, text: str) -> str:
        cache_key = text
        if cache_key in self._cache:
            return self._cache[cache_key]

        masked, tokens = _mask_placeholders(text)
        translated = self._translator.translate(masked)
        if translated is None:
            translated = text
        translated = _unmask_placeholders(str(translated), tokens)
        self._cache[cache_key] = translated
        return translated


def _build_translator(
    backend: str,
    source_language: LanguageSpec,
    target_language: LanguageSpec,
) -> Tuple[Optional[GoogleTranslatorBackend], str, List[str]]:
    warnings: List[str] = []
    if source_language.code == target_language.code:
        return None, "same-language", warnings

    normalized_backend = (backend or "auto").strip().lower()
    if normalized_backend == "none":
        return None, "none", warnings

    if normalized_backend in {"auto", "google"}:
        try:
            return (
                GoogleTranslatorBackend(source_language.code, target_language.code),
                "google",
                warnings,
            )
        except Exception as exc:
            warnings.append(
                "Runtime translator unavailable; continuing with AI translation map only "
                f"({exc})."
            )
            return None, "none", warnings

    warnings.append(f"Unsupported translator backend '{backend}'. Using AI translation map only.")
    return None, "none", warnings


def translate_sheet(
    sheet,
    sheet_name: str,
    language: LanguageSpec,
    source_language: LanguageSpec,
    mode: str,
    dry_run: bool,
    translator_backend: str,
    translation_map: Dict[str, Dict[str, str]],
) -> Dict[str, object]:
    result = {
        "sheet": sheet_name,
        "headers_added": [],
        "headers_renamed": [],
        "header_conflicts": [],
        "merged_values": 0,
        "cells_translated": 0,
        "cells_copied_media": 0,
        "cells_translated_ai_map": 0,
        "cells_translated_runtime": 0,
        "cells_pending_translation": 0,
        "cells_skipped_existing": 0,
        "cells_missing_source": 0,
        "runtime_backend": "none",
        "warnings": [],
        "errors": [],
    }

    header_state = ensure_language_columns(
        sheet=sheet,
        sheet_name=sheet_name,
        language=language,
        dry_run=dry_run or mode == "validate-only",
    )
    target_columns = header_state["target_columns"]
    result["headers_added"] = list(header_state["added_headers"])
    result["headers_renamed"] = list(header_state["renamed_headers"])
    result["header_conflicts"] = list(header_state["header_conflicts"])
    result["merged_values"] = int(header_state["merged_values"])

    if mode in {"validate-only", "header-sync"}:
        return result

    entries = _collect_headers(sheet)
    source_columns: Dict[str, Optional[int]] = {}
    for category in CATEGORY_ORDER:
        source_columns[category] = _choose_source_column(
            sheet=sheet,
            entries=entries,
            category=category,
            source_language=source_language,
        )

    translator, runtime_backend, backend_warnings = _build_translator(
        backend=translator_backend,
        source_language=source_language,
        target_language=language,
    )
    result["runtime_backend"] = runtime_backend
    result["warnings"].extend(backend_warnings)

    map_entries = _translation_map_entry_count(translation_map)
    if runtime_backend == "none" and source_language.code != language.code and map_entries == 0:
        result["warnings"].append(
            f"{sheet_name}: no AI translation map provided and runtime translator unavailable. "
            "Cross-language text cells will remain pending."
        )

    memory = _build_translation_memory(sheet, source_columns, target_columns) if mode == "add-missing" else {}
    warning_cap = 25

    for row_idx in range(2, sheet.max_row + 1):
        for category in CATEGORY_ORDER:
            src_col = source_columns.get(category)
            dst_col = target_columns.get(category)
            if not src_col or not dst_col:
                continue

            src_val = sheet.cell(row=row_idx, column=src_col).value
            if not _cell_has_value(src_val):
                result["cells_missing_source"] += 1
                continue

            dst_val = sheet.cell(row=row_idx, column=dst_col).value
            if mode == "add-missing" and _cell_has_value(dst_val):
                result["cells_skipped_existing"] += 1
                continue

            if category in MEDIA_CATEGORIES:
                new_value = src_val
                translation_source = "media-copy"
                if _cell_has_value(new_value):
                    result["cells_copied_media"] += 1
            else:
                src_text = str(src_val)
                memory_key = (category, src_text.strip())
                if memory_key in memory:
                    new_value = memory[memory_key]
                    translation_source = "memory"
                else:
                    map_translation = _lookup_translation_map(translation_map, category, src_text)
                    if map_translation is not None:
                        new_value = map_translation
                        translation_source = "ai-map"
                    elif translator is not None:
                        try:
                            new_value = translator.translate_text(src_text)
                            translation_source = "runtime"
                        except Exception as exc:
                            result["errors"].append(
                                f"{sheet_name}: row {row_idx}, {category}: translation failed ({exc})"
                            )
                            continue
                    elif source_language.code == language.code:
                        new_value = src_text
                        translation_source = "same-language"
                    else:
                        result["cells_pending_translation"] += 1
                        if len(result["warnings"]) < warning_cap:
                            result["warnings"].append(
                                f"{sheet_name}: row {row_idx}, {category} pending translation."
                            )
                        continue

                    if not _placeholders_match(src_text, str(new_value)):
                        result["cells_pending_translation"] += 1
                        if len(result["warnings"]) < warning_cap:
                            result["warnings"].append(
                                f"{sheet_name}: row {row_idx}, {category} skipped due placeholder mismatch."
                            )
                        continue
                    memory[memory_key] = str(new_value)

            if str(dst_val or "") == str(new_value or ""):
                continue

            if not dry_run:
                sheet.cell(row=row_idx, column=dst_col, value=new_value)
            result["cells_translated"] += 1
            if translation_source == "ai-map":
                result["cells_translated_ai_map"] += 1
            elif translation_source == "runtime":
                result["cells_translated_runtime"] += 1

    return result


def _resolve_xlsform_path(file_arg: Optional[str]) -> Path:
    if file_arg:
        return Path(file_arg).resolve()
    if CONFIG_AVAILABLE:
        try:
            config = ProjectConfig()
            return config.get_full_xlsform_path().resolve()
        except Exception:
            pass
    return Path("survey.xlsx").resolve()


def _get_project_dir_for_logging(xlsx_path: Path) -> Path:
    if CONFIG_AVAILABLE:
        try:
            config = ProjectConfig()
            return config.project_dir
        except Exception:
            pass
    return xlsx_path.parent


def ensure_default_language(
    workbook,
    language: LanguageSpec,
    set_mode: str,
    dry_run: bool,
) -> Dict[str, object]:
    changes = {
        "header_added": False,
        "status": "no_change",
        "value": None,
        "warnings": [],
    }

    if set_mode == "never":
        return changes

    sheet = workbook["settings"] if "settings" in workbook.sheetnames else workbook.create_sheet("settings")

    header_map: Dict[str, int] = {}
    max_col = max(sheet.max_column, 1)
    for col in range(1, max_col + 1):
        cell_val = sheet.cell(row=1, column=col).value
        if _cell_has_value(cell_val):
            header_map[normalize_header_name(str(cell_val))] = col

    default_col = header_map.get("default_language")
    if not default_col:
        default_col = max_col + 1
        if not dry_run:
            sheet.cell(row=1, column=default_col, value="default_language")
        changes["header_added"] = True

    desired = language.header_label
    current_val = sheet.cell(row=2, column=default_col).value
    current_text = str(current_val).strip() if _cell_has_value(current_val) else ""

    should_set = False
    if set_mode == "always":
        should_set = True
        changes["status"] = "updated" if current_text else "set"
    elif set_mode == "auto":
        if not current_text:
            should_set = True
            changes["status"] = "set"
        else:
            changes["status"] = "no_change"
    else:
        changes["warnings"].append(f"Unknown set-default-language mode: {set_mode}")

    if should_set:
        if not dry_run:
            sheet.cell(row=2, column=default_col, value=desired)
        changes["value"] = desired
    else:
        changes["value"] = current_text or None

    return changes


def build_report(
    xlsx_path: Path,
    intent: str,
    mode: str,
    source_language: LanguageSpec,
    target_languages: List[LanguageSpec],
    sheet_results: Dict[str, Dict[str, object]],
    settings_changes: Dict[str, object],
    snapshot_revision: str,
    dry_run: bool,
    saved: bool,
    runtime_backend: str,
    ai_map_entries: int,
    extra_warnings: Optional[List[str]] = None,
) -> Dict[str, object]:
    warnings: List[str] = []
    errors: List[str] = []
    header_conflicts: List[str] = []

    summary = {
        "headers_added": 0,
        "headers_renamed": 0,
        "merged_values": 0,
        "cells_translated": 0,
        "cells_copied_media": 0,
        "cells_translated_ai_map": 0,
        "cells_translated_runtime": 0,
        "cells_pending_translation": 0,
        "cells_skipped_existing": 0,
        "cells_missing_source": 0,
    }

    header_changes = {}
    for sheet_name, payload in sheet_results.items():
        added_headers = payload.get("headers_added", [])
        renamed_headers = payload.get("headers_renamed", [])
        sheet_conflicts = payload.get("header_conflicts", [])

        header_changes[sheet_name] = {
            "added_headers": added_headers,
            "renamed_headers": renamed_headers,
        }
        summary["headers_added"] += len(added_headers)
        summary["headers_renamed"] += len(renamed_headers)
        summary["merged_values"] += int(payload.get("merged_values", 0))
        summary["cells_translated"] += int(payload.get("cells_translated", 0))
        summary["cells_copied_media"] += int(payload.get("cells_copied_media", 0))
        summary["cells_translated_ai_map"] += int(payload.get("cells_translated_ai_map", 0))
        summary["cells_translated_runtime"] += int(payload.get("cells_translated_runtime", 0))
        summary["cells_pending_translation"] += int(payload.get("cells_pending_translation", 0))
        summary["cells_skipped_existing"] += int(payload.get("cells_skipped_existing", 0))
        summary["cells_missing_source"] += int(payload.get("cells_missing_source", 0))

        warnings.extend(payload.get("warnings", []))
        errors.extend(payload.get("errors", []))
        header_conflicts.extend(sheet_conflicts)

    warnings.extend(settings_changes.get("warnings", []))
    if extra_warnings:
        warnings.extend(extra_warnings)
    if summary["cells_pending_translation"] > 0:
        warnings.append(
            "Some target cells are still pending translation. Re-run with an AI translation map "
            "or enable runtime fallback."
        )

    # Keep output concise: deduplicate repeated warnings/errors while preserving order.
    warnings = list(dict.fromkeys(warnings))
    errors = list(dict.fromkeys(errors))
    header_conflicts = list(dict.fromkeys(header_conflicts))

    status = "completed"
    if errors:
        status = "failed"
    elif dry_run:
        status = "dry_run"
    elif warnings or header_conflicts:
        status = "completed_with_warnings"

    report = {
        "status": status,
        "intent": intent,
        "mode": mode,
        "file": str(xlsx_path),
        "source_language": source_language.header_label,
        "target_languages": [lang.header_label for lang in target_languages],
        "summary": summary,
        "translation_runtime": {
            "backend": runtime_backend,
            "ai_map_entries": ai_map_entries,
        },
        "header_changes": header_changes,
        "header_conflicts": header_conflicts,
        "settings_changes": {
            "default_language": settings_changes.get("status", "no_change"),
            "header_added": settings_changes.get("header_added", False),
            "value": settings_changes.get("value"),
        },
        "warnings": warnings,
        "errors": errors,
        "rollback": {"snapshot_revision": snapshot_revision or None},
        "saved": bool(saved),
    }
    return report


def print_structured_report(report: Dict[str, object]) -> None:
    print("# XLSFORM_TRANSLATION_RESULT")
    print(f"status: {report['status']}")
    print(f"intent: {report['intent']}")
    print(f"mode: {report['mode']}")
    print(f"file: {report['file']}")
    print(f"source_language: {report['source_language']}")
    print("target_languages:")
    for item in report["target_languages"]:
        print(f"  - {item}")
    print("summary:")
    for key, value in report["summary"].items():
        print(f"  {key}: {value}")
    print("translation_runtime:")
    print(f"  backend: {report['translation_runtime']['backend']}")
    print(f"  ai_map_entries: {report['translation_runtime']['ai_map_entries']}")
    print("settings_changes:")
    print(f"  default_language: {report['settings_changes']['default_language']}")
    print(f"  header_added: {str(report['settings_changes']['header_added']).lower()}")
    if report["settings_changes"]["value"] is not None:
        print(f"  value: {report['settings_changes']['value']}")
    print("rollback:")
    print(f"  snapshot_revision: {report['rollback']['snapshot_revision']}")

    print("header_conflicts:")
    if report["header_conflicts"]:
        for item in report["header_conflicts"]:
            print(f"  - {item}")
    else:
        print("  - none")

    print("warnings:")
    if report["warnings"]:
        for item in report["warnings"]:
            print(f"  - {item}")
    else:
        print("  - none")

    print("errors:")
    if report["errors"]:
        for item in report["errors"]:
            print(f"  - {item}")
    else:
        print("  - none")


def log_translation_activity(report: Dict[str, object], project_dir: Path) -> None:
    if not LOGGING_AVAILABLE:
        return
    try:
        logger = ActivityLogger(project_dir=project_dir)
        details = (
            f"Mode: {report['mode']}\n"
            f"Source: {report['source_language']}\n"
            f"Targets: {', '.join(report['target_languages'])}\n"
            f"Headers added: {report['summary']['headers_added']}\n"
            f"Headers renamed: {report['summary']['headers_renamed']}\n"
            f"Cells translated: {report['summary']['cells_translated']}\n"
            f"Media copied: {report['summary']['cells_copied_media']}\n"
            f"Translated (AI map): {report['summary']['cells_translated_ai_map']}\n"
            f"Translated (runtime): {report['summary']['cells_translated_runtime']}\n"
            f"Pending translations: {report['summary']['cells_pending_translation']}\n"
            f"Runtime backend: {report['translation_runtime']['backend']}\n"
            f"AI map entries: {report['translation_runtime']['ai_map_entries']}\n"
            f"Snapshot revision: {report['rollback']['snapshot_revision']}"
        )
        logger.log_action(
            action_type="translate",
            description=f"Translation command {report['status']}",
            details=details,
        )
    except Exception:
        pass


def _load_translation_map_from_args(
    inline_json: Optional[str],
    json_file: Optional[str],
) -> Tuple[Dict[str, Dict[str, str]], List[str], List[str]]:
    warnings: List[str] = []
    errors: List[str] = []
    raw_payload: Optional[Dict[str, Any]] = None

    if inline_json and json_file:
        errors.append("Use either --translation-map or --translation-map-file, not both.")
        return {"*": {}}, warnings, errors

    if inline_json:
        try:
            payload = json.loads(inline_json)
        except Exception as exc:
            errors.append(f"Failed to parse --translation-map JSON: {exc}")
            return {"*": {}}, warnings, errors
        if not isinstance(payload, dict):
            errors.append("--translation-map must be a JSON object.")
            return {"*": {}}, warnings, errors
        raw_payload = payload

    if json_file:
        file_path = Path(json_file).expanduser().resolve()
        if not file_path.exists():
            errors.append(f"Translation map file not found: {file_path}")
            return {"*": {}}, warnings, errors
        try:
            payload = json.loads(file_path.read_text(encoding="utf-8"))
        except Exception as exc:
            errors.append(f"Failed to parse translation map file '{file_path}': {exc}")
            return {"*": {}}, warnings, errors
        if not isinstance(payload, dict):
            errors.append(f"Translation map file must contain a JSON object: {file_path}")
            return {"*": {}}, warnings, errors
        raw_payload = payload

    prepared, prep_warnings = _prepare_translation_map(raw_payload)
    warnings.extend(prep_warnings)
    return prepared, warnings, errors


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Translate XLSForm columns with natural language intent parsing.")
    parser.add_argument(
        "instruction",
        nargs="*",
        help="Natural language command text, e.g. 'add Bangla language'",
    )
    parser.add_argument("--file", "-f", help="Path to XLSForm file (default uses project config).")
    parser.add_argument("--language", "-l", action="append", help="Target language token (repeatable).")
    parser.add_argument("--languages", help="Comma-separated target languages.")
    parser.add_argument("--source-language", "--source", help="Source language token (default: settings/default English).")
    parser.add_argument(
        "--mode",
        choices=["add-missing", "update-all", "validate-only", "header-sync"],
        help="Execution mode override.",
    )
    parser.add_argument(
        "--translator",
        choices=["auto", "google", "none"],
        default="auto",
        help="Runtime translation backend. 'auto' uses google if available, otherwise AI map only.",
    )
    parser.add_argument(
        "--translation-map",
        help="Inline JSON map of source text to translated text (AI-generated contextual translations).",
    )
    parser.add_argument(
        "--translation-map-file",
        help="Path to JSON file with AI-generated contextual translations.",
    )
    parser.add_argument(
        "--set-default-language",
        choices=["auto", "always", "never"],
        default="auto",
        help="How to manage settings.default_language.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Preview changes without saving the workbook.")
    parser.add_argument("--json", action="store_true", help="Print JSON output instead of structured text.")
    return parser.parse_args()


def main() -> int:
    args = parse_arguments()

    instruction_text = " ".join(args.instruction).strip()
    parsed_intent = parse_instruction_intent(instruction_text)

    intent = str(parsed_intent["intent"])
    mode = args.mode or str(parsed_intent["mode"])

    target_tokens: List[str] = []
    if args.languages:
        target_tokens.extend(_split_language_tokens(args.languages))
    if args.language:
        for token in args.language:
            target_tokens.extend(_split_language_tokens(token))
    if parsed_intent["language_tokens"]:
        target_tokens.extend(parsed_intent["language_tokens"])

    dedup_tokens: List[str] = []
    seen = set()
    for token in target_tokens:
        key = _normalize_alias(token)
        if key and key not in seen:
            seen.add(key)
            dedup_tokens.append(token)
    target_tokens = dedup_tokens

    if not target_tokens:
        print("Error: No target language detected. Provide e.g. '/xlsform-translate add Bangla language'.")
        return 1

    target_languages: List[LanguageSpec] = []
    language_errors: List[str] = []
    for token in target_tokens:
        lang, err = resolve_language_token(token, allow_unknown_code=False)
        if err:
            language_errors.append(err)
            continue
        target_languages.append(lang)

    if language_errors:
        for err in language_errors:
            print(f"Error: {err}")
        return 1

    translation_map, map_warnings, map_errors = _load_translation_map_from_args(
        inline_json=args.translation_map,
        json_file=args.translation_map_file,
    )
    if map_errors:
        for err in map_errors:
            print(f"Error: {err}")
        return 1
    ai_map_entries = _translation_map_entry_count(translation_map)

    xlsx_path = _resolve_xlsform_path(args.file)
    if not xlsx_path.exists():
        print(f"Error: XLSForm file not found: {xlsx_path}")
        return 1

    history_manager = None
    lock_acquired = False
    snapshot_revision = ""
    saved = False

    try:
        if not args.dry_run and HISTORY_AVAILABLE:
            history_manager = WorkbookHistoryManager(xlsform_path=xlsx_path, project_dir=xlsx_path.parent)
            history_manager.acquire_lock()
            lock_acquired = True

        workbook = openpyxl.load_workbook(xlsx_path)

        if args.source_language:
            source_language, source_err = resolve_language_token(args.source_language, allow_unknown_code=False)
            if source_err:
                print(f"Error: {source_err}")
                return 1
        else:
            source_language = None
            if "settings" in workbook.sheetnames:
                settings_sheet = workbook["settings"]
                header_map = {}
                for col in range(1, max(settings_sheet.max_column, 1) + 1):
                    val = settings_sheet.cell(row=1, column=col).value
                    if _cell_has_value(val):
                        header_map[normalize_header_name(str(val))] = col
                default_col = header_map.get("default_language")
                if default_col:
                    existing = settings_sheet.cell(row=2, column=default_col).value
                    if _cell_has_value(existing):
                        source_language, _ = resolve_language_token(str(existing), allow_unknown_code=True)

            if not source_language:
                source_language = LanguageSpec(code="en", display_name="English")

        effective_dry_run = args.dry_run or mode == "validate-only"

        sheet_results: Dict[str, Dict[str, object]] = {}
        runtime_backends = set()
        for sheet_name in ["survey", "choices"]:
            if sheet_name not in workbook.sheetnames:
                sheet_results[sheet_name] = {
                    "sheet": sheet_name,
                    "headers_added": [],
                    "headers_renamed": [],
                    "header_conflicts": [],
                    "merged_values": 0,
                    "cells_translated": 0,
                    "cells_copied_media": 0,
                    "cells_translated_ai_map": 0,
                    "cells_translated_runtime": 0,
                    "cells_pending_translation": 0,
                    "cells_skipped_existing": 0,
                    "cells_missing_source": 0,
                    "runtime_backend": "none",
                    "warnings": [f"Sheet '{sheet_name}' not found, skipped."],
                    "errors": [],
                }
                continue

            sheet = workbook[sheet_name]
            aggregate = {
                "sheet": sheet_name,
                "headers_added": [],
                "headers_renamed": [],
                "header_conflicts": [],
                "merged_values": 0,
                "cells_translated": 0,
                "cells_copied_media": 0,
                "cells_translated_ai_map": 0,
                "cells_translated_runtime": 0,
                "cells_pending_translation": 0,
                "cells_skipped_existing": 0,
                "cells_missing_source": 0,
                "runtime_backend": "none",
                "warnings": [],
                "errors": [],
            }

            for target_language in target_languages:
                translated = translate_sheet(
                    sheet=sheet,
                    sheet_name=sheet_name,
                    language=target_language,
                    source_language=source_language,
                    mode=mode,
                    dry_run=effective_dry_run,
                    translator_backend=args.translator,
                    translation_map=translation_map,
                )
                runtime_backends.add(str(translated.get("runtime_backend", "none")))
                aggregate["headers_added"].extend(translated["headers_added"])
                aggregate["headers_renamed"].extend(translated["headers_renamed"])
                aggregate["header_conflicts"].extend(translated["header_conflicts"])
                aggregate["merged_values"] += int(translated["merged_values"])
                aggregate["cells_translated"] += int(translated["cells_translated"])
                aggregate["cells_copied_media"] += int(translated["cells_copied_media"])
                aggregate["cells_translated_ai_map"] += int(translated["cells_translated_ai_map"])
                aggregate["cells_translated_runtime"] += int(translated["cells_translated_runtime"])
                aggregate["cells_pending_translation"] += int(translated["cells_pending_translation"])
                aggregate["cells_skipped_existing"] += int(translated["cells_skipped_existing"])
                aggregate["cells_missing_source"] += int(translated["cells_missing_source"])
                aggregate["warnings"].extend(translated["warnings"])
                aggregate["errors"].extend(translated["errors"])

            sheet_results[sheet_name] = aggregate

        if "google" in runtime_backends:
            report_runtime_backend = "google"
        elif len(runtime_backends) == 1:
            report_runtime_backend = next(iter(runtime_backends))
        elif runtime_backends:
            report_runtime_backend = ",".join(sorted(runtime_backends))
        else:
            report_runtime_backend = "none"

        settings_changes = ensure_default_language(
            workbook=workbook,
            language=target_languages[0],
            set_mode=args.set_default_language,
            dry_run=effective_dry_run,
        )

        pre_report = build_report(
            xlsx_path=xlsx_path.resolve(),
            intent=intent,
            mode=mode,
            source_language=source_language,
            target_languages=target_languages,
            sheet_results=sheet_results,
            settings_changes=settings_changes,
            snapshot_revision=snapshot_revision,
            dry_run=effective_dry_run,
            saved=False,
            runtime_backend=report_runtime_backend,
            ai_map_entries=ai_map_entries,
            extra_warnings=map_warnings,
        )

        if not effective_dry_run and pre_report["status"] != "failed":
            total_changes = (
                pre_report["summary"]["headers_added"]
                + pre_report["summary"]["headers_renamed"]
                + pre_report["summary"]["cells_translated"]
                + pre_report["summary"]["merged_values"]
            )
            if settings_changes.get("status") in {"set", "updated"}:
                total_changes += 1
            if settings_changes.get("header_added"):
                total_changes += 1

            if total_changes > 0:
                if history_manager is not None:
                    snapshot = history_manager.create_snapshot(
                        action_type="translate",
                        description=(
                            f"Pre-change snapshot before translation to "
                            f"{', '.join([lang.header_label for lang in target_languages])}"
                        ),
                        details=f"Mode: {mode}",
                        command="/xlsform-translate",
                    )
                    snapshot_revision = snapshot.get("revision_id", "")

                workbook.save(xlsx_path)
                saved = True

        final_report = build_report(
            xlsx_path=xlsx_path.resolve(),
            intent=intent,
            mode=mode,
            source_language=source_language,
            target_languages=target_languages,
            sheet_results=sheet_results,
            settings_changes=settings_changes,
            snapshot_revision=snapshot_revision,
            dry_run=effective_dry_run,
            saved=saved,
            runtime_backend=report_runtime_backend,
            ai_map_entries=ai_map_entries,
            extra_warnings=map_warnings,
        )

        if saved and final_report["status"] != "failed":
            log_translation_activity(final_report, project_dir=_get_project_dir_for_logging(xlsx_path))

        if args.json:
            print(json.dumps(final_report, ensure_ascii=False, indent=2))
        else:
            print_structured_report(final_report)

        return 1 if final_report["status"] == "failed" else 0
    finally:
        if lock_acquired and history_manager is not None:
            history_manager.release_lock()


if __name__ == "__main__":
    raise SystemExit(main())
