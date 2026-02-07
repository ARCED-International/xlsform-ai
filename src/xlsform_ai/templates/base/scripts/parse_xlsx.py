#!/usr/bin/env python3
"""
Excel Question Parser for XLSForm AI.

Extracts questions from generic Excel worksheets and emits JSON.
Usage: python parse_xlsx.py <file.xlsx> --sheet "Questions"
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional

# Add scripts directory to Python path for sibling imports.
_scripts_dir = Path(__file__).parent.resolve()
if str(_scripts_dir) not in sys.path:
    sys.path.insert(0, str(_scripts_dir))

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

try:
    from form_structure import normalize_header_name
except Exception:
    def normalize_header_name(value: str) -> str:
        return str(value).strip().lower().replace(" ", "_")


TEXT_HEADER_CANDIDATES = [
    "question",
    "question_text",
    "text",
    "label",
    "prompt",
]
TYPE_HEADER_CANDIDATES = [
    "type",
    "question_type",
]
NAME_HEADER_CANDIDATES = [
    "name",
    "variable",
    "variable_name",
    "field_name",
    "code",
]
CHOICES_HEADER_CANDIDATES = [
    "choices",
    "options",
    "option",
    "responses",
    "answer_options",
]
CONSTRAINT_HEADER_CANDIDATES = [
    "constraint",
    "validation",
    "rule",
]
REQUIRED_HEADER_CANDIDATES = [
    "required",
    "is_required",
    "mandatory",
]


def _header_map(sheet) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    max_col = max(sheet.max_column, 1)
    for col in range(1, max_col + 1):
        val = sheet.cell(row=1, column=col).value
        if val is None:
            continue
        key = normalize_header_name(str(val))
        if key:
            mapping[key] = col
    return mapping


def _find_col(mapping: Dict[str, int], candidates: List[str]) -> Optional[int]:
    for candidate in candidates:
        key = normalize_header_name(candidate)
        if key in mapping:
            return mapping[key]
    return None


def _parse_required(value) -> bool:
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"yes", "y", "true", "1", "required"}


def _choice_value(label: str) -> str:
    value = re.sub(r"\s+", "_", label.strip().lower())
    value = re.sub(r"[^a-z0-9_]", "", value)
    return value[:48] or "option"


def _parse_choices(raw_value) -> Optional[List[Dict[str, str]]]:
    if raw_value is None:
        return None
    text = str(raw_value).strip()
    if not text:
        return None

    parts = [p.strip() for p in re.split(r"[,\n;|]", text) if p.strip()]
    if not parts:
        return None
    return [{"value": _choice_value(label), "label": label} for label in parts]


def _infer_type(question_text: str, choices: Optional[List[Dict[str, str]]]) -> str:
    if choices:
        if "select all" in question_text.lower() or "multiple" in question_text.lower():
            return "select_multiple"
        return "select_one"

    text = question_text.lower()
    if any(token in text for token in ["how old", "age", "how many", "number of"]):
        return "integer"
    if any(token in text for token in ["date", "when"]):
        return "date"
    if any(token in text for token in ["location", "gps", "coordinate"]):
        return "geopoint"
    return "text"


def extract_questions_from_xlsx(xlsx_path: str | Path, sheet: Optional[str] = None) -> List[Dict[str, object]]:
    """Extract questions from an Excel worksheet."""
    source = Path(xlsx_path).resolve()
    workbook = openpyxl.load_workbook(source, data_only=True)
    try:
        if sheet:
            if sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet not found: '{sheet}'")
            ws = workbook[sheet]
        else:
            ws = workbook[workbook.sheetnames[0]]

        mapping = _header_map(ws)
        text_col = _find_col(mapping, TEXT_HEADER_CANDIDATES)
        type_col = _find_col(mapping, TYPE_HEADER_CANDIDATES)
        name_col = _find_col(mapping, NAME_HEADER_CANDIDATES)
        choices_col = _find_col(mapping, CHOICES_HEADER_CANDIDATES)
        constraint_col = _find_col(mapping, CONSTRAINT_HEADER_CANDIDATES)
        required_col = _find_col(mapping, REQUIRED_HEADER_CANDIDATES)

        if text_col is None:
            raise ValueError(
                "Could not find a question text column. Expected one of: "
                + ", ".join(TEXT_HEADER_CANDIDATES)
            )

        questions: List[Dict[str, object]] = []
        question_num = 0

        for row_idx in range(2, ws.max_row + 1):
            text_value = ws.cell(row=row_idx, column=text_col).value
            if text_value is None or not str(text_value).strip():
                continue

            question_num += 1
            question_text = str(text_value).strip()

            parsed_choices = _parse_choices(ws.cell(row=row_idx, column=choices_col).value) if choices_col else None
            explicit_type = (
                str(ws.cell(row=row_idx, column=type_col).value).strip()
                if type_col and ws.cell(row=row_idx, column=type_col).value is not None
                else ""
            )
            q_type = explicit_type if explicit_type else _infer_type(question_text, parsed_choices)

            question: Dict[str, object] = {
                "number": str(question_num),
                "text": question_text,
                "type": q_type,
                "choices": parsed_choices,
                "required": _parse_required(ws.cell(row=row_idx, column=required_col).value) if required_col else False,
            }

            if name_col and ws.cell(row=row_idx, column=name_col).value is not None:
                question["name"] = str(ws.cell(row=row_idx, column=name_col).value).strip()
            if constraint_col and ws.cell(row=row_idx, column=constraint_col).value is not None:
                constraint = str(ws.cell(row=row_idx, column=constraint_col).value).strip()
                if constraint:
                    question["constraint"] = constraint

            questions.append(question)

        return questions
    finally:
        workbook.close()


def extract_questions(xlsx_path: str | Path, pages: Optional[str] = None, sheet: Optional[str] = None):
    """Compatibility alias used in older command protocols."""
    _ = pages
    return extract_questions_from_xlsx(xlsx_path, sheet=sheet)


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract questions from Excel file for XLSForm import")
    parser.add_argument("xlsx_file", help="Path to Excel (.xlsx) file")
    parser.add_argument("--sheet", help="Sheet name to parse (default: first sheet)", default=None)
    parser.add_argument("--output", help="Output JSON file", default=None)
    args = parser.parse_args()

    source_path = Path(args.xlsx_file).resolve()
    if not source_path.exists():
        print(f"Error: File not found: {source_path}")
        sys.exit(1)

    try:
        questions = extract_questions_from_xlsx(source_path, sheet=args.sheet)
    except Exception as exc:
        print(f"Error: Failed to parse Excel file: {exc}")
        sys.exit(1)

    try:
        from log_activity import ActivityLogger

        logger = ActivityLogger()
        question_summary = ", ".join(
            [f"{q.get('text', 'Unknown')[:30]} ({q.get('type', 'unknown')})" for q in questions[:5]]
        )
        if len(questions) > 5:
            question_summary += f"... and {len(questions) - 5} more"

        logger.log_action(
            action_type="import_xlsx",
            description=f"Imported {len(questions)} questions from Excel",
            details=f"Source: {source_path}\nSheet: {args.sheet or '(first sheet)'}\nQuestions: {question_summary}",
        )
    except Exception:
        pass

    output = {
        "source": str(source_path),
        "sheet": args.sheet,
        "count": len(questions),
        "questions": questions,
    }
    json_output = json.dumps(output, ensure_ascii=False, indent=2)

    if args.output:
        output_path = Path(args.output).resolve()
        output_path.write_text(json_output, encoding="utf-8")
        print(f"Saved {len(questions)} questions to {output_path}")
    else:
        print(json_output)


if __name__ == "__main__":
    main()
