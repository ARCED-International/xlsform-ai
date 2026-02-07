#!/usr/bin/env python3
"""
XLSForm Validator

Runs structured validation using:
1) Local structural checks (openpyxl)
2) XLSForm -> XForm conversion (pyxform)
3) Offline ODK Validate jar (if available)

Usage:
  python validate_form.py survey.xlsx
  python validate_form.py survey.xlsx --json
  python validate_form.py survey.xlsx --jar-path tools/ODK-Validate.jar
"""

import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# Ensure UTF-8 encoding for Windows console output
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except AttributeError:
        pass

# Add scripts directory for sibling imports whether run from project root or scripts dir
_SCRIPTS_DIR = Path(__file__).parent.resolve()
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

_PROJECT_ROOT = _SCRIPTS_DIR.parent

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def _cell_has_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _normalize_header_name(header_name: str) -> str:
    try:
        from form_structure import normalize_header_name

        return normalize_header_name(header_name)
    except Exception:
        return header_name.strip().lower()


def _build_header_map(sheet) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for idx, cell in enumerate(sheet[1]):
        if _cell_has_value(cell.value):
            key = _normalize_header_name(str(cell.value).strip())
            headers[key] = idx
    return headers


def validate_choices_sheet(sheet) -> Tuple[List[str], List[str], set]:
    errors: List[str] = []
    warnings: List[str] = []
    choice_lists: set = set()

    headers = _build_header_map(sheet)

    required = ["list_name", "name", "label"]
    for col in required:
        if col not in headers:
            errors.append(f"Missing required column in choices sheet: '{col}'")
    if errors:
        return errors, warnings, choice_lists

    list_name_col = headers["list_name"]
    name_col = headers["name"]

    choice_names: Dict[str, Dict[str, int]] = {}
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) <= max(list_name_col, name_col):
            continue
        list_name_val = row[list_name_col]
        name_val = row[name_col]
        if not (_cell_has_value(list_name_val) and _cell_has_value(name_val)):
            continue

        list_name = str(list_name_val).strip()
        name = str(name_val).strip()
        choice_lists.add(list_name)

        if list_name not in choice_names:
            choice_names[list_name] = {}

        if name in choice_names[list_name]:
            first_row = choice_names[list_name][name]
            errors.append(
                f"Duplicate choice name '{name}' in list '{list_name}' at rows {first_row} and {row_idx}"
            )
        else:
            choice_names[list_name][name] = row_idx

    return errors, warnings, choice_lists


def validate_survey_sheet(sheet, choice_lists: set) -> Tuple[List[str], List[str]]:
    errors: List[str] = []
    warnings: List[str] = []

    headers = _build_header_map(sheet)

    required = ["type", "name", "label"]
    for col in required:
        if col not in headers:
            errors.append(f"Missing required column in survey sheet: '{col}'")
    if errors:
        return errors, warnings

    type_col = headers["type"]
    name_col = headers["name"]
    relevant_col = headers.get("relevant")
    calculation_col = headers.get("calculation")

    names: Dict[str, int] = {}
    group_stack: List[Tuple[str, int]] = []
    repeat_stack: List[Tuple[str, int]] = []

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row_values = list(row)

        if len(row_values) > name_col and _cell_has_value(row_values[name_col]):
            name = str(row_values[name_col]).strip()
            if name in names:
                errors.append(f"Duplicate question name '{name}' at rows {names[name]} and {row_idx}")
            else:
                names[name] = row_idx

        question_type = ""
        if len(row_values) > type_col and _cell_has_value(row_values[type_col]):
            question_type = str(row_values[type_col]).strip()
        if not question_type:
            continue

        normalized_type = question_type.lower()
        if normalized_type.startswith("begin group") or normalized_type == "begin_group":
            group_stack.append((question_type, row_idx))
        elif normalized_type.startswith("end group") or normalized_type == "end_group":
            if not group_stack:
                errors.append(f"Unmatched end_group at row {row_idx}")
            else:
                group_stack.pop()
        elif normalized_type.startswith("begin repeat") or normalized_type == "begin_repeat":
            repeat_stack.append((question_type, row_idx))
        elif normalized_type.startswith("end repeat") or normalized_type == "end_repeat":
            if not repeat_stack:
                errors.append(f"Unmatched end_repeat at row {row_idx}")
            else:
                repeat_stack.pop()

        if normalized_type.startswith("select_one ") or normalized_type.startswith("select_multiple "):
            parts = question_type.split()
            if len(parts) < 2:
                errors.append(f"Invalid select question type syntax at row {row_idx}: '{question_type}'")
            else:
                list_name = parts[1].strip()
                if list_name and list_name not in choice_lists:
                    errors.append(
                        f"Missing choice list '{list_name}' referenced at survey row {row_idx}"
                    )

        if relevant_col is not None and len(row_values) > relevant_col:
            relevant_val = row_values[relevant_col]
            if _cell_has_value(relevant_val):
                text = str(relevant_val).strip()
                if "$" in text and "${" not in text:
                    warnings.append(f"Possible malformed relevant expression at row {row_idx}: '{text}'")

        if calculation_col is not None and len(row_values) > calculation_col:
            calculation_val = row_values[calculation_col]
            if _cell_has_value(calculation_val):
                text = str(calculation_val).strip()
                if "$" in text and "${" not in text:
                    warnings.append(f"Possible malformed calculation expression at row {row_idx}: '{text}'")

    for _, row_idx in group_stack:
        errors.append(f"Unclosed begin_group at row {row_idx}")
    for _, row_idx in repeat_stack:
        errors.append(f"Unclosed begin_repeat at row {row_idx}")

    return errors, warnings


def validate_xlsxform(xlsx_path: Path) -> Tuple[List[str], List[str], List[str]]:
    """Run local XLSForm checks."""
    errors: List[str] = []
    warnings: List[str] = []
    suggestions: List[str] = []

    try:
        workbook = openpyxl.load_workbook(xlsx_path)

        if "survey" not in workbook.sheetnames:
            errors.append("Missing required sheet: 'survey'")
        if "choices" not in workbook.sheetnames:
            warnings.append("Missing optional sheet: 'choices'")

        choice_lists = set()
        if "choices" in workbook.sheetnames:
            choices_errors, choices_warnings, choice_lists = validate_choices_sheet(workbook["choices"])
            errors.extend(choices_errors)
            warnings.extend(choices_warnings)

        if "survey" in workbook.sheetnames:
            survey_errors, survey_warnings = validate_survey_sheet(workbook["survey"], choice_lists)
            errors.extend(survey_errors)
            warnings.extend(survey_warnings)

        workbook.close()
    except Exception as exc:
        errors.append(f"Failed to load workbook: {exc}")

    return errors, warnings, suggestions


def validate_form(xlsx_path: Path) -> Tuple[List[str], List[str], List[str]]:
    """Compatibility wrapper used by agent docs."""
    return validate_xlsxform(Path(xlsx_path))


def _find_odk_validate_jar(explicit_jar_path: Optional[str]) -> Optional[Path]:
    candidates: List[Path] = []
    if explicit_jar_path:
        candidates.append(Path(explicit_jar_path))

    env_jar = os.getenv("ODK_VALIDATE_JAR")
    if env_jar:
        candidates.append(Path(env_jar))

    candidates.extend(
        [
            _PROJECT_ROOT / "tools" / "ODK-Validate.jar",
            Path.cwd() / "tools" / "ODK-Validate.jar",
        ]
    )

    for candidate in candidates:
        if candidate.exists():
            return candidate.resolve()
    return None


def _parse_odk_output(output_text: str) -> Tuple[List[str], List[str], List[str]]:
    errors: List[str] = []
    warnings: List[str] = []
    info: List[str] = []

    for raw_line in output_text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        normalized = line.lower()
        if "error" in normalized:
            errors.append(line)
        elif "warning" in normalized:
            warnings.append(line)
        else:
            info.append(line)

    return errors, warnings, info


def _convert_xlsform_to_xform(xlsform_path: Path, xform_path: Path) -> Tuple[bool, List[str], List[str]]:
    """Convert XLSForm to XForm XML using pyxform."""
    errors: List[str] = []
    warnings: List[str] = []

    try:
        from pyxform.xls2xform import xls2xform_convert
    except ImportError:
        errors.append("pyxform is not installed. Install with: pip install pyxform")
        return False, errors, warnings
    except Exception as exc:
        errors.append(f"Failed to load pyxform: {exc}")
        return False, errors, warnings

    try:
        xls2xform_convert(str(xlsform_path), str(xform_path))
    except TypeError:
        try:
            xls2xform_convert(str(xlsform_path), str(xform_path), warnings=warnings)
        except Exception as exc:
            errors.append(f"XLSForm -> XForm conversion failed: {exc}")
            return False, errors, warnings
    except Exception as exc:
        errors.append(f"XLSForm -> XForm conversion failed: {exc}")
        return False, errors, warnings

    if not xform_path.exists():
        errors.append("XLSForm -> XForm conversion did not produce XML output.")
        return False, errors, warnings

    return True, errors, warnings


def run_odk_validate(
    form_path: Path,
    jar_path: Optional[Path],
    timeout_seconds: int = 180,
) -> Dict[str, object]:
    result: Dict[str, object] = {
        "enabled": True,
        "ran": False,
        "status": "skipped",
        "jar_path": str(jar_path) if jar_path else None,
        "exit_code": None,
        "errors": [],
        "warnings": [],
        "info": [],
        "command": None,
        "xform_path": None,
    }

    if jar_path is None:
        result["status"] = "jar_not_found"
        result["warnings"] = [
            "Offline ODK validation skipped because tools/ODK-Validate.jar was not found."
        ]
        return result

    form_suffix = form_path.suffix.lower()
    xform_input_path: Optional[Path] = None

    if form_suffix == ".xml":
        xform_input_path = form_path
    elif form_suffix in [".xlsx", ".xlsm", ".xls"]:
        temp_dir_path = Path(tempfile.mkdtemp(prefix="xlsform-ai-validate-"))
        try:
            xform_candidate = temp_dir_path / f"{form_path.stem}.xml"
            ok, convert_errors, convert_warnings = _convert_xlsform_to_xform(form_path, xform_candidate)
            if not ok:
                pyxform_missing = any("pyxform is not installed" in item for item in convert_errors)
                result["status"] = "pyxform_not_found" if pyxform_missing else "xform_conversion_failed"
                if pyxform_missing:
                    result["warnings"] = convert_errors
                else:
                    result["errors"] = convert_errors
                if convert_warnings:
                    result["warnings"] = list(result["warnings"]) + convert_warnings
                return result

            result["xform_path"] = str(xform_candidate)
            command = ["java", "-jar", str(jar_path), str(xform_candidate)]
            result["command"] = " ".join(command)

            try:
                proc = subprocess.run(
                    command,
                    capture_output=True,
                    text=True,
                    timeout=timeout_seconds,
                    check=False,
                )
            except FileNotFoundError:
                result["status"] = "java_not_found"
                result["warnings"] = [
                    "Java runtime not found. Install Java to enable offline ODK validation."
                ]
                return result
            except subprocess.TimeoutExpired:
                result["status"] = "timeout"
                result["warnings"] = [f"ODK validation timed out after {timeout_seconds} seconds."]
                return result
            except Exception as exc:
                result["status"] = "execution_error"
                result["warnings"] = [f"ODK validation failed to start: {exc}"]
                return result

            combined_output = "\n".join(
                part for part in [proc.stdout.strip(), proc.stderr.strip()] if part
            )
            parsed_errors, parsed_warnings, parsed_info = _parse_odk_output(combined_output)

            result["ran"] = True
            result["status"] = "completed"
            result["exit_code"] = proc.returncode
            result["errors"] = parsed_errors
            result["warnings"] = list(convert_warnings) + parsed_warnings
            result["info"] = parsed_info

            if proc.returncode != 0 and not parsed_errors:
                result["warnings"] = list(convert_warnings) + list(parsed_warnings) + [
                    f"ODK Validate exited with code {proc.returncode}."
                ]

            return result
        finally:
            try:
                shutil.rmtree(temp_dir_path, ignore_errors=True)
            except Exception:
                pass
    else:
        result["status"] = "unsupported_input"
        result["warnings"] = [
            f"Unsupported input format '{form_suffix}'. Provide .xlsx or .xml for ODK validation."
        ]
        return result

    if xform_input_path is None:
        result["status"] = "xform_conversion_failed"
        result["errors"] = ["XForm input could not be prepared for ODK validation."]
        return result

    result["xform_path"] = str(xform_input_path.resolve())
    command = ["java", "-jar", str(jar_path), str(xform_input_path)]
    result["command"] = " ".join(command)

    try:
        proc = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=timeout_seconds,
            check=False,
        )
    except FileNotFoundError:
        result["status"] = "java_not_found"
        result["warnings"] = [
            "Java runtime not found. Install Java to enable offline ODK validation."
        ]
        return result
    except subprocess.TimeoutExpired:
        result["status"] = "timeout"
        result["warnings"] = [f"ODK validation timed out after {timeout_seconds} seconds."]
        return result
    except Exception as exc:
        result["status"] = "execution_error"
        result["warnings"] = [f"ODK validation failed to start: {exc}"]
        return result

    combined_output = "\n".join(
        part for part in [proc.stdout.strip(), proc.stderr.strip()] if part
    )
    parsed_errors, parsed_warnings, parsed_info = _parse_odk_output(combined_output)

    result["ran"] = True
    result["status"] = "completed"
    result["exit_code"] = proc.returncode
    result["errors"] = parsed_errors
    result["warnings"] = parsed_warnings
    result["info"] = parsed_info

    if proc.returncode != 0 and not parsed_errors:
        result["warnings"] = list(parsed_warnings) + [
            f"ODK Validate exited with code {proc.returncode}."
        ]

    return result


def build_validation_report(
    xlsx_path: Path,
    local_errors: List[str],
    local_warnings: List[str],
    local_suggestions: List[str],
    odk_result: Dict[str, object],
) -> Dict[str, object]:
    combined_errors = list(local_errors)
    combined_warnings = list(local_warnings)

    if odk_result.get("enabled", True):
        for item in odk_result.get("errors", []):
            combined_errors.append(f"[odk] {item}")
        for item in odk_result.get("warnings", []):
            combined_warnings.append(f"[odk] {item}")

    report = {
        "file": str(xlsx_path.resolve()),
        "timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "valid": len(combined_errors) == 0,
        "summary": {
            "error_count": len(combined_errors),
            "warning_count": len(combined_warnings),
            "suggestion_count": len(local_suggestions),
        },
        "engines": {
            "local": {
                "status": "passed" if len(local_errors) == 0 else "failed",
                "error_count": len(local_errors),
                "warning_count": len(local_warnings),
            },
            "odk_validate": {
                "status": odk_result.get("status"),
                "ran": bool(odk_result.get("ran")),
                "jar_path": odk_result.get("jar_path"),
                "xform_path": odk_result.get("xform_path"),
                "exit_code": odk_result.get("exit_code"),
                "error_count": len(odk_result.get("errors", [])),
                "warning_count": len(odk_result.get("warnings", [])),
            },
        },
        "errors": combined_errors,
        "warnings": combined_warnings,
        "suggestions": local_suggestions,
        "details": {
            "local": {
                "errors": local_errors,
                "warnings": local_warnings,
                "suggestions": local_suggestions,
            },
            "odk_validate": {
                "errors": odk_result.get("errors", []),
                "warnings": odk_result.get("warnings", []),
                "info": odk_result.get("info", []),
                "command": odk_result.get("command"),
                "xform_path": odk_result.get("xform_path"),
            },
        },
    }
    return report


def print_structured_report(report: Dict[str, object]) -> None:
    summary = report["summary"]
    engines = report["engines"]
    local_engine = engines["local"]
    odk_engine = engines["odk_validate"]

    print("# XLSFORM_VALIDATION_RESULT")
    print(f"valid: {str(report['valid']).lower()}")
    print(f"file: {report['file']}")
    print(f"timestamp_utc: {report['timestamp_utc']}")
    print("summary:")
    print(f"  errors: {summary['error_count']}")
    print(f"  warnings: {summary['warning_count']}")
    print(f"  suggestions: {summary['suggestion_count']}")
    print("engines:")
    print(f"  local.status: {local_engine['status']}")
    print(f"  local.errors: {local_engine['error_count']}")
    print(f"  local.warnings: {local_engine['warning_count']}")
    print(f"  odk_validate.status: {odk_engine['status']}")
    print(f"  odk_validate.ran: {str(odk_engine['ran']).lower()}")
    if odk_engine.get("jar_path"):
        print(f"  odk_validate.jar: {odk_engine['jar_path']}")
    if odk_engine.get("xform_path"):
        print(f"  odk_validate.xform: {odk_engine['xform_path']}")
    if odk_engine.get("exit_code") is not None:
        print(f"  odk_validate.exit_code: {odk_engine['exit_code']}")

    print("errors:")
    if report["errors"]:
        for item in report["errors"]:
            print(f"  - {item}")
    else:
        print("  - none")

    print("warnings:")
    if report["warnings"]:
        for item in report["warnings"]:
            print(f"  - {item}")
    else:
        print("  - none")

    print("suggestions:")
    if report["suggestions"]:
        for item in report["suggestions"]:
            print(f"  - {item}")
    else:
        print("  - none")


def log_validation_activity(report: Dict[str, object]) -> None:
    try:
        from log_activity import ActivityLogger

        logger = ActivityLogger()
        logger.log_action(
            action_type="validate",
            description=f"Form validation {'passed' if report['valid'] else 'failed'}",
            details=(
                f"Errors: {report['summary']['error_count']}\n"
                f"Warnings: {report['summary']['warning_count']}\n"
                f"Suggestions: {report['summary']['suggestion_count']}\n"
                f"ODK status: {report['engines']['odk_validate']['status']}"
            ),
        )
    except Exception:
        pass


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate XLSForm (.xlsx) or XForm (.xml)")
    parser.add_argument("form_file", nargs="?", default="survey.xlsx", help="Path to XLSForm/XForm file")
    parser.add_argument("--jar-path", help="Path to ODK-Validate.jar (default: tools/ODK-Validate.jar)")
    parser.add_argument("--timeout", type=int, default=180, help="ODK validator timeout in seconds")
    parser.add_argument("--skip-odk", action="store_true", help="Run local checks only")
    parser.add_argument("--structured", action="store_true", help="Print structured report (default)")
    parser.add_argument("--json", action="store_true", help="Print full report as JSON")

    args = parser.parse_args()

    form_path = Path(args.form_file).resolve()
    if not form_path.exists():
        print(f"Error: File not found: {form_path}")
        sys.exit(1)

    if form_path.suffix.lower() in [".xlsx", ".xlsm", ".xls"]:
        local_errors, local_warnings, local_suggestions = validate_xlsxform(form_path)
    else:
        local_errors = []
        local_warnings = [f"Local XLSForm checks skipped for {form_path.suffix or 'unknown'} input."]
        local_suggestions = []

    if args.skip_odk:
        odk_result = {
            "enabled": False,
            "ran": False,
            "status": "disabled",
            "jar_path": None,
            "xform_path": None,
            "exit_code": None,
            "errors": [],
            "warnings": [],
            "info": [],
            "command": None,
        }
    else:
        jar_path = _find_odk_validate_jar(args.jar_path)
        odk_result = run_odk_validate(form_path, jar_path, timeout_seconds=args.timeout)

    report = build_validation_report(
        xlsx_path=form_path,
        local_errors=local_errors,
        local_warnings=local_warnings,
        local_suggestions=local_suggestions,
        odk_result=odk_result,
    )

    log_validation_activity(report)

    if args.json:
        print(json.dumps(report, indent=2))
    else:
        print_structured_report(report)

    sys.exit(1 if not report["valid"] else 0)


if __name__ == "__main__":
    main()
